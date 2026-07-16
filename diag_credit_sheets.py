"""
diag_credit_sheets.py  —  Paste output back to Claude.
Checks Debt, Capital Structure, and Credit Rating sheets for AEE + YORW.
"""
import os, glob
import openpyxl

REPORTS_DIR = r"E:\PowerAcademy\data\reports"

TARGET_SHEETS = [
    'Debt Summary (Reported)',
    'Debt Analysis',
    'Capital Structure Summary',
    'Capital Structure Details',
    'Credit Ratings',
    'Current Ratings',
    'Ratings History',
    'Credit Ratios (x)',
]

files = sorted(glob.glob(os.path.join(REPORTS_DIR, '*.xlsx')))
samples = {}
for f in files:
    b = os.path.basename(f)
    for t in ['AEE', 'YORW', 'D']:
        import re
        m = re.search(r'(?:NASDAQGS|NASDAQGM|NASDAQ|NYSE)' + t + r'_Report', b)
        if m and t not in samples:
            samples[t] = f

for ticker, fpath in samples.items():
    print(f"\n{'='*70}")
    print(f"TICKER: {ticker}  FILE: {os.path.basename(fpath)}")
    print('='*70)
    wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)

    for sheet_name in TARGET_SHEETS:
        if sheet_name not in wb.sheetnames:
            print(f"\n  ['{sheet_name}' NOT FOUND]")
            continue
        ws = wb[sheet_name]
        print(f"\n--- '{sheet_name}' (first 40 non-empty rows) ---")
        count = 0
        for row in ws.iter_rows(values_only=True):
            vals = [str(v)[:55] if v is not None else '' for v in row]
            if any(v.strip() for v in vals):
                print(f"  R{count+1}: {vals[:8]}")
                count += 1
            if count >= 40:
                print("  ...")
                break
    wb.close()