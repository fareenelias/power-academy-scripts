"""
explore_exec_sheets.py  —  Run this first, paste output back to Claude.
Reads one CapIQ Excel file and prints all sheet names + first few rows
of any sheet that looks like it contains executive/board/bio data.
"""

import os, glob
import openpyxl

REPORTS_DIR = r"E:\PowerAcademy\data\reports"

files = sorted(glob.glob(os.path.join(REPORTS_DIR, "*.xlsx")))
if not files:
    print("No .xlsx files found in", REPORTS_DIR)
    exit(1)

# Use first file as sample
sample = files[0]
print(f"Sample file: {os.path.basename(sample)}")
print()

wb = openpyxl.load_workbook(sample, read_only=True, data_only=True)
print(f"All sheets ({len(wb.sheetnames)}):")
for s in wb.sheetnames:
    print(f"  '{s}'")
print()

# Keywords that suggest exec/board content
EXEC_KEYWORDS = ['exec', 'officer', 'director', 'board', 'manage', 'bio',
                 'leadership', 'people', 'govern', 'comp', 'remun']

for sheet_name in wb.sheetnames:
    if any(kw in sheet_name.lower() for kw in EXEC_KEYWORDS):
        ws = wb[sheet_name]
        print(f"=== Sheet: '{sheet_name}' ===")
        rows = list(ws.iter_rows(values_only=True))
        print(f"  Dimensions: {ws.dimensions}, rows sampled: {min(len(rows), 20)}")
        for i, row in enumerate(rows[:20]):
            # Show non-empty rows only
            vals = [str(v)[:60] if v is not None else '' for v in row]
            if any(v for v in vals):
                print(f"  R{i+1}: {vals}")
        print()

wb.close()

print("=" * 60)
print("Now spot-checking a water utility file for comparison...")
water_files = [f for f in files if any(t in os.path.basename(f).upper()
               for t in ['MSEX','AWK','CWT','AWR','WTRG','HTO','YORW','GWRS'])]
if water_files:
    wb2 = openpyxl.load_workbook(water_files[0], read_only=True, data_only=True)
    print(f"\nWater file: {os.path.basename(water_files[0])}")
    print("Sheets:", wb2.sheetnames)
    for sheet_name in wb2.sheetnames:
        if any(kw in sheet_name.lower() for kw in EXEC_KEYWORDS):
            ws = wb2[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            print(f"\n  Sheet '{sheet_name}' first 15 rows:")
            for i, row in enumerate(rows[:15]):
                vals = [str(v)[:60] if v is not None else '' for v in row]
                if any(v for v in vals):
                    print(f"    R{i+1}: {vals}")
    wb2.close()