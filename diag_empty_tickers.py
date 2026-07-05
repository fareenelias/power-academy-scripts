"""
extract_executives.py  —  July 5 2026
Reads CapIQ Excel 'People' and 'Compensation' sheets for all 23 companies.
Outputs: E:\PowerAcademy\data\executives_export.json

People sheet layout (confirmed):
  Top Executives section   — bio at col 11
  Board of Directors section — bio at col 14, tenure at col 7, begin_year at col 5
Compensation sheet: per-exec salary + total comp for most recent fiscal year.

Run:  python extract_executives.py
"""

import json, os, glob, re
import openpyxl

REPORTS_DIR = r"E:\PowerAcademy\data\reports"
OUTPUT_FILE  = r"E:\PowerAcademy\data\executives_export.json"

# Companies where CapIQ ticker differs from our internal ticker
# (HTO/XIFR are now correctly named in filenames — map kept for safety only)
TICKER_MAP = {}

def get_ticker(filename):
    # Must match NASDAQGS / NASDAQGM / NASDAQCM before bare NASDAQ
    m = re.search(r'(?:NYSE|NASDAQGS|NASDAQGM|NASDAQCM|NASDAQ|NasdaqGS|NasdaqCM|OTC)([A-Z0-9]+)_Report',
                  os.path.basename(filename))
    if not m:
        return None
    t = m.group(1)
    return TICKER_MAP.get(t, t)

def clean(v):
    """Return stripped string or None for NA / empty."""
    if v is None:
        return None
    s = str(v).strip()
    return s if s and s.upper() not in ('NA', 'N/A', 'NONE', '') else None

def to_int(v):
    if v is None:
        return None
    try:
        n = int(str(v).strip())
        return n if n > 0 else None
    except Exception:
        return None

def comp_val(v):
    """Parse a compensation cell to int dollars, or None."""
    if v is None:
        return None
    s = str(v).strip()
    if s.upper() in ('NA', 'N/A', 'NONE', ''):
        return None
    try:
        return int(float(s.replace(',', '')))
    except Exception:
        return None

def norm_name(name):
    """Normalise for compensation matching: strip nicknames, suffixes, lowercase."""
    s = re.sub(r'"[^"]*"', '', str(name))           # remove "Marty" style nicknames
    s = re.sub(r'\s+(Jr\.?|Sr\.?|II|III|IV)\s*$', '', s, flags=re.I)
    s = re.sub(r'\xa0', ' ', s)                      # non-breaking spaces
    return re.sub(r'\s+', ' ', s).strip().lower()

# ── People sheet ─────────────────────────────────────────────────────────────

def parse_people(ws):
    """
    Returns (management_list, board_list).
    Section boundaries detected by sentinel strings in col 0.
    """
    management, board = [], []
    mode = None           # 'exec' | 'board'
    awaiting_exec_hdr  = False
    awaiting_board_hdr = False

    for row in ws.iter_rows(values_only=True):
        col0 = clean(row[0]) if row else None

        # ── Section sentinels ────────────────────────────────────────────────
        if col0 == 'Top Executives':
            awaiting_exec_hdr = True
            continue
        if col0 == 'Board of Directors':
            awaiting_exec_hdr  = False
            awaiting_board_hdr = True
            mode = None
            continue
        # Sections that signal we're past the data we want
        if col0 in ('Other Board Members', 'Professionals',
                    'Takeover Defenses', 'Latest Holders - Insiders/Strategic Owners'):
            break
        if col0 == 'Average':
            mode = None
            continue

        # ── Column header rows ───────────────────────────────────────────────
        if col0 == 'Name':
            if awaiting_exec_hdr:
                mode = 'exec'
                awaiting_exec_hdr = False
            elif awaiting_board_hdr:
                mode = 'board'
                awaiting_board_hdr = False
            continue

        # ── Data rows ────────────────────────────────────────────────────────
        if not col0:
            continue

        if mode == 'exec':
            management.append({
                'name':   clean(row[0]),
                'title':  clean(row[1]),
                'age':    to_int(row[3]),
                'gender': clean(row[4]),
                'email':  clean(row[7]),
                'bio':    clean(row[11]) if len(row) > 11 else None,
            })

        elif mode == 'board':
            board.append({
                'name':       clean(row[0]),
                'title':      clean(row[1]),
                'age':        to_int(row[3]),
                'gender':     clean(row[4]),
                'begin_year': to_int(row[5]),
                'end_year':   to_int(row[6]),
                'tenure':     clean(row[7]),
                'email':      clean(row[10]),
                'bio':        clean(row[14]) if len(row) > 14 else None,
            })

    return management, board


# ── Compensation sheet ────────────────────────────────────────────────────────

def parse_compensation(ws):
    """
    Returns {normalised_name: {salary, total_comp}} for the most recent fiscal year.
    The "As Reported Total Executive Compensation" row is used for total_comp.
    """
    result      = {}
    in_summary  = False
    recent_col  = None
    current_key = None

    for row in ws.iter_rows(values_only=True):
        col0 = str(row[0]).strip() if row[0] is not None else ''

        # ── Find the summary section ─────────────────────────────────────────
        if col0 == 'Executive Compensation - Summary':
            in_summary = True
            continue
        if not in_summary:
            continue

        # ── Period header row: blank col0, 'FY' in other cols ───────────────
        if not col0 and any(v and 'FY' in str(v) for v in row[1:]):
            period_cols = [i for i, v in enumerate(row) if v and 'FY' in str(v)]
            if period_cols:
                recent_col = period_cols[-1]
            continue

        # ── Executive block header: "Name - Title" with all other cols empty ─
        if (col0 and ' - ' in col0
                and all((row[i] is None or str(row[i]).strip() == '')
                        for i in range(1, min(4, len(row))))):
            raw_name    = col0.split(' - ')[0].strip()
            current_key = norm_name(raw_name)
            if current_key not in result:
                result[current_key] = {}
            continue

        # ── Data rows ────────────────────────────────────────────────────────
        if current_key is None or recent_col is None:
            continue

        v = row[recent_col] if recent_col < len(row) else None
        n = comp_val(v)

        if col0 == 'Salary':
            result[current_key]['salary'] = n
        elif col0 == 'As Reported Total Executive Compensation':
            result[current_key]['total_comp'] = n

    return result


# ── Name matching ─────────────────────────────────────────────────────────────

def link_comp(management, comp):
    """Attach compensation data to management list using exact then last-name fallback."""
    for person in management:
        key = norm_name(person['name'])
        c   = comp.get(key)

        # Fallback: match on last name if exact key not found
        if not c:
            last = key.split()[-1]
            for ck, cv in comp.items():
                if ck.split()[-1] == last and abs(len(ck) - len(key)) < 10:
                    c = cv
                    break

        person['salary_latest']     = c.get('salary')     if c else None
        person['total_comp_latest'] = c.get('total_comp') if c else None


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    files = sorted(glob.glob(os.path.join(REPORTS_DIR, '*.xlsx')))
    print(f'Found {len(files)} Excel files\n')

    output = {}

    for fpath in files:
        ticker = get_ticker(fpath)
        if not ticker:
            print(f'  SKIP (no ticker): {os.path.basename(fpath)}')
            continue

        print(f'  {ticker:<6}', end=' ')
        try:
            wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)

            management, board = [], []
            comp = {}

            if 'People' in wb.sheetnames:
                management, board = parse_people(wb['People'])

            if 'Compensation' in wb.sheetnames:
                comp = parse_compensation(wb['Compensation'])

            wb.close()

            link_comp(management, comp)

            output[ticker] = {
                'management': management,
                'board':      board,
                'source':     'S&P Capital IQ',
                'as_of':      '06-28-2026',
            }

            print(f'{len(management):2} mgmt  {len(board):2} board')

        except Exception as e:
            print(f'ERROR — {e}')

    with open(OUTPUT_FILE, 'w', encoding='utf-8') as f:
        json.dump(output, f, indent=2, ensure_ascii=False)

    print(f'\n[✓] Saved {len(output)} tickers → {OUTPUT_FILE}')

    print('\nSummary:')
    for t, d in sorted(output.items()):
        mg = d['management']
        bd = d['board']
        comp_pct = sum(1 for p in mg if p.get('total_comp_latest')) / max(len(mg), 1)
        print(f'  {t:<6}  {len(mg):2}mgmt  {len(bd):2}board  '
              f'{comp_pct*100:.0f}% have comp data')

if __name__ == '__main__':
    main()