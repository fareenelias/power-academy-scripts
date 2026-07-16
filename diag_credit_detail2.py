"""
diag_credit_detail2.py  —  Paste output back to Claude.
Shows ALL rows of Current Ratings + Debt Analysis + full columns of Capital
Structure Details for EIX (has multiple rated entities: SCE) and D.
"""
import os, glob, re
import openpyxl

REPORTS_DIR = r"E:\PowerAcademy\data\reports"

files = sorted(glob.glob(os.path.join(REPORTS_DIR, '*.xlsx')))
samples = {}
for f in files:
    b = os.path.basename(f)
    for t in ['EIX', 'D']:
        m = re.search(r'(?:NASDAQGS|NASDAQGM|NASDAQ|NYSE)' + t + r'_Report', b)
        if m and t not in samples:
            samples[t] = f

for ticker, fpath in samples.items():
    print(f"\n{'='*70}")
    print(f"TICKER: {ticker}")
    print('='*70)
    wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)

    # ── Current Ratings: ALL rows (no limit) ─────────────────────────────────
    if 'Current Ratings' in wb.sheetnames:
        ws = wb['Current Ratings']
        print("\n--- Current Ratings (ALL non-empty rows) ---")
        count = 0
        for row in ws.iter_rows(values_only=True):
            vals = [str(v)[:55] if v is not None else '' for v in row]
            if any(v.strip() for v in vals):
                print(f"  R{count+1}: {vals[:8]}")
                count += 1
    else:
        print("\n  [Current Ratings NOT FOUND]")

    # ── Debt Analysis: structure around Coverage Ratios section ──────────────
    if 'Debt Analysis' in wb.sheetnames:
        ws = wb['Debt Analysis']
        print("\n--- Debt Analysis (rows 1-80, looking for Coverage Ratios) ---")
        count = 0
        for row in ws.iter_rows(max_row=80, values_only=True):
            vals = [str(v)[:55] if v is not None else '' for v in row]
            if any(v.strip() for v in vals):
                print(f"  R{count+1}: {vals[:7]}")
            count += 1
    else:
        print("\n  [Debt Analysis NOT FOUND]")

    # ── Capital Structure Details: show all columns for first 5 data rows ─────
    if 'Capital Structure Details' in wb.sheetnames:
        ws = wb['Capital Structure Details']
        print("\n--- Capital Structure Details: FULL columns for first 5 data rows ---")
        header_seen = False
        data_count  = 0
        row_count   = 0
        for row in ws.iter_rows(values_only=True):
            row_count += 1
            col0 = str(row[0]).strip() if row[0] else ''
            if col0 == 'Capital Structure Description':
                vals = [(i, str(v)[:40]) for i, v in enumerate(row) if v is not None]
                print(f"  HEADER cols: {vals}")
                header_seen = True
                continue
            if not header_seen: continue
            if col0.startswith('Total') or col0.startswith('*') or not col0: continue
            vals = [(i, str(v)[:40]) for i, v in enumerate(row) if v is not None]
            print(f"  R{row_count}: {vals}")
            data_count += 1
            if data_count >= 5: break
    else:
        print("\n  [Capital Structure Details NOT FOUND]")

    wb.close()