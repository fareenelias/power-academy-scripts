#!/usr/bin/env python3
# -*- coding: utf-8 -*-
r"""
extract_peers.py
================

Build `data\peer_universe.json` -- the Tier 2 lightweight comp profile for the
large-cap and mid-cap utilities OUTSIDE the 24-name coverage universe.

Why this is a separate script and not a flag on extract_all.py
--------------------------------------------------------------
`extract_all.py` reads EVERY .xlsx in `data\reports\`. Dropping peer workbooks in
that folder would fold 25 more companies into `capiq_export.json` and
`executives_export.json`, take the dashboard company list from 24 to 49, and
silently invalidate every "23/24 names" health check. It would present as a data
problem, not a routing one -- the same shape as the two rippers coexisting.

So: peers get their own folder, their own script, their own output file. This
script writes `peer_universe.json` and NOTHING else. There is a hard path guard.

One system, not two
-------------------
The workbook parsing is NOT reimplemented here. `Extractor` comes from
`extract_aqn_financials.py` and the rate-case / ratings / M&A parsers come from
`extract_all.py`. Two copies of the same parser drift, and on 2026-07-30 a stale
container copy of `assemble.py` overwrote a patched device copy and dropped a
metric on the very next run. Import, never copy.

Usage
-----
    python extract_peers.py --peer-reports E:\PowerAcademy\data\peer_reports \
                            --out          E:\PowerAcademy\data\peer_universe.json \
                           [--precedents   E:\PowerAcademy\data\precedents.json] \
                           [--dry-run] [--allow-loss]

Guarantees
----------
  * MERGE, NEVER REBUILD. Existing peers are preserved; a peer whose workbook is
    absent this run is carried through untouched.
  * NO SILENT LOSS. A key-level diff runs before the write. If any peer loses a
    populated field the write is REFUSED unless --allow-loss is passed and the
    losses are printed. Four separate data-loss incidents in this project were
    caught only by diffing keys before and after, and every merge reported
    success either way.
  * NEVER INVENT A NUMBER. Every value is read out of a named cell of a named
    sheet, or computed from such values and labelled `_basis`. A field the
    workbook does not print is absent, with a reason in `_missing`.
  * ATOMIC WRITE. Temp file -> re-parse -> os.replace.
"""

from __future__ import annotations

import argparse
import collections
import importlib.util
import json
import os
import re
import sys

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl is required:  pip install openpyxl --break-system-packages")


# ---------------------------------------------------------------------------
# guards
# ---------------------------------------------------------------------------

PROTECTED_OUTPUTS = {
    "capiq_export.json", "executives_export.json", "broker_research.json",
    "earnings_calls.json", "precedents.json", "market_data.json",
    "moodys_credit.json", "sp_credit.json", "fitch_credit.json",
}

# The 24-name coverage universe. A peer workbook for any of these is a routing
# error -- almost certainly --peer-reports was pointed at data\reports\.
COVERAGE_UNIVERSE = {
    "NEE", "D", "ETR", "CMS", "PPL", "AEE", "POR", "EIX", "PCG", "HE", "EVRG",
    "ES", "VST", "TLN", "XIFR", "AWR", "CWT", "YORW", "GWRS", "AWK", "WTRG",
    "HTO", "MSEX", "AQN",
}

# Asset class uses the precedents.json vocabulary so a peer can be matched to a
# comp set without a translation layer. Judgment call per name, stated openly
# rather than inferred from a CapIQ SIC code that lumps everything as "electric".
ASSET_CLASS = {
    "SRE": "integrated", "SO": "integrated", "DUK": "integrated",
    "XEL": "integrated", "PEG": "integrated", "DTE": "integrated",
    "WEC": "integrated", "LNT": "integrated", "PNW": "electric",
    "EXC": "electric", "AEP": "electric", "ED": "electric", "FE": "electric",
    "OGE": "electric", "IDA": "electric", "AVA": "integrated",
    "BKH": "integrated", "NWE": "integrated", "TXNM": "electric",
    "CEG": "ipp", "NRG": "ipp",
    "ATO": "gas_ldc", "NI": "gas_ldc", "CNP": "gas_ldc", "OGS": "gas_ldc",
    "SWX": "gas_ldc", "SR": "gas_ldc", "NJR": "gas_ldc", "UGI": "gas_ldc",
    "CPK": "gas_ldc",
    "FTS": "integrated", "EMA": "integrated", "H": "electric",
}

# Reported in a currency other than USD. Listed explicitly so a mixed-currency
# average can never happen by accident -- the AQN lesson, where one house
# quoting C$ against eight quoting US$ overstated consensus by ~40%.
NON_USD = {"FTS": "CAD", "EMA": "CAD", "H": "CAD"}


def die(msg):
    sys.exit("REFUSED: " + msg)


def check_guards(peer_dir, out_path):
    out_base = os.path.basename(out_path).lower()
    if out_base in PROTECTED_OUTPUTS:
        die("--out is %s. This script writes peer_universe.json and nothing else."
            % out_base)
    if not out_base.startswith("peer"):
        die("--out must be a peer_* file (got %r). Refusing to write outside the "
            "peer namespace." % out_base)
    if not os.path.isdir(peer_dir):
        die("--peer-reports %r is not a directory. Create it and put the peer "
            "CapIQ workbooks there -- NOT in data\\reports\\." % peer_dir)


# ---------------------------------------------------------------------------
# import the existing parsers (one system, not two)
# ---------------------------------------------------------------------------

def load_module(path, name):
    if not os.path.isfile(path):
        die("cannot find %s at %r. Pass --scripts-dir." % (name, path))
    spec = importlib.util.spec_from_file_location(name, path)
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    return mod


# ---------------------------------------------------------------------------
# small helpers
# ---------------------------------------------------------------------------

NULLISH = {"", "NA", "N/A", "NM", "-", "--", "NULL", "NONE"}


def num(v):
    if v is None:
        return None
    s = str(v).strip().replace(",", "").replace("$", "").replace("%", "")
    if s.upper() in NULLISH:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def norm(v):
    return re.sub(r"\s+", " ", str(v or "")).strip()


def ticker_of(fname):
    """CapIQ names files <LegalName><EXCHANGE><TICKER>_Report_<MM-DD-YYYY>.xlsx."""
    base = re.sub(r"_Report_.*$", "", os.path.splitext(os.path.basename(fname))[0])
    for pre in ("NASDAQGS", "NASDAQGM", "NASDAQCM", "NASDAQ", "NYSEARCA",
                "NYSEAMERICAN", "NYSEMKT", "NYSE", "TSXV", "TSX", "AMEX", "OTCPK"):
        if pre in base:
            t = base.split(pre)[-1]
            if t:
                return t.upper(), pre
    return None, None


def vintage_of(fname):
    m = re.search(r"_Report_(\d{2})-(\d{2})-(\d{4})", os.path.basename(fname))
    return "%s-%s-%s" % (m.group(3), m.group(1), m.group(2)) if m else None


def rows_of(wb, sheet):
    if sheet not in wb.sheetnames:
        return []
    return [tuple(r) for r in wb[sheet].iter_rows(values_only=True)]


def find(rows, label, col=0, start=0, exact=False):
    lab = label.lower()
    for i in range(start, len(rows)):
        cellv = norm(rows[i][col] if col < len(rows[i]) else None).lower()
        if (cellv == lab) if exact else (cellv.startswith(lab)):
            return i
    return None


# ---------------------------------------------------------------------------
# peer-specific reads: valuation multiples, market cap, allowed ROE, ratings
# ---------------------------------------------------------------------------

def read_valuation(wb, missing):
    """Key Stats 'Valuation Multiples' block.

    These are RATIOS -- numerator and denominator are in the same currency -- so
    unlike `tev` and `market_cap` they are directly comparable across a
    mixed-currency peer set with no FX conversion. That is why they are stored
    for Canadian names where the absolute EV is not.
    """
    ks = rows_of(wb, "Key Stats")
    if not ks:
        missing["valuation"] = "no 'Key Stats' sheet"
        return {}
    hdr = find(ks, "period ended", start=(find(ks, "valuation multiples") or 0))
    if hdr is None:
        missing["valuation"] = "no valuation-multiples header on Key Stats"
        return {}
    labels = [norm(v) for v in ks[hdr - 1]] if hdr else []

    # FY1 = the first forward estimate column, i.e. the first header ending 'E'.
    fy1 = None
    for c, lab in enumerate(labels):
        if lab and lab.rstrip().upper().endswith("E"):
            fy1 = c
            break
    if fy1 is None:
        missing["valuation"] = "no forward-estimate ('...E') column on Key Stats"
        return {}

    out = {"_fy1_period": labels[fy1]}
    for key, label in (("ev_ebitda_fy1", "TEV/ EBITDA"),
                       ("ev_revenue_fy1", "TEV/ Total Revenue"),
                       ("pe_fy1", "Price/ EPS"),
                       ("p_book", "Price/ Book")):
        r = find(ks, label.lower(), start=hdr)
        v = num(ks[r][fy1]) if r is not None and fy1 < len(ks[r]) else None
        if v is None:
            missing[key] = "row %r absent or NM in the %s column" % (label, labels[fy1])
        else:
            out[key] = v
    return out


def read_capitalisation(wb, missing):
    """Key Stats 'Latest Capitalization'. Currency comes from the LABEL.

    CapIQ prices a dual-listed issuer off its PRIMARY listing, so the labels read
    'Market Capitalization (C$000)' inside a workbook whose statements are US$000.
    The currency is therefore parsed out of the printed label, never assumed from
    the workbook's reported-currency code.
    """
    ks = rows_of(wb, "Key Stats")
    start = find(ks, "latest capitalization")
    if start is None:
        missing["market_cap"] = "no 'Latest Capitalization' block on Key Stats"
        return {}
    out = {}
    for key, label in (("market_cap_000", "market capitalization"),
                       ("close_price", "closing price"),
                       ("shares_out", "common shares outstanding"),
                       ("tev_000", "total enterprise value")):
        r = find(ks, label, start=start)
        if r is None:
            missing[key] = "no %r row" % label
            continue
        printed = norm(ks[r][0])
        # Take the currency out of the parenthesised group that actually CONTAINS a
        # currency symbol. Scanning for the first "(" instead reads "(TEV)" out of
        # "Total Enterprise Value (TEV) (C$000)" and silently labels a C$ figure
        # with a nonsense currency -- which is worse than no label at all.
        cur = None
        for grp in re.findall(r"\(([^)]*)\)", printed):
            # Every alternative carries an explicit currency SYMBOL. Allowing a
            # bare 3-letter code here matches "(TEV)" and labels a C$ figure as
            # currency "TEV".
            m = re.match(r"\s*(C\$|US\$|A\$|[A-Z]{2,3}\$|\$|€|£)\s*0{0,3}\s*$", grp)
            if m:
                cur = {"C$": "CAD", "US$": "USD", "$": "USD",
                       "A$": "AUD"}.get(m.group(1), m.group(1).rstrip("$"))
                break
        v = None
        for c in range(1, min(9, len(ks[r]))):
            v = num(ks[r][c])
            if v is not None:
                break
        if v is None:
            missing[key] = "row %r present but no numeric value" % printed
            continue
        out[key] = v
        out[key + "_label"] = printed
        # A share count is not money. Only emit a currency where the label
        # actually printed one -- defaulting to USD would make an unlabelled
        # figure look confirmed.
        if cur:
            out[key + "_currency"] = cur
        elif key != "shares_out":
            missing[key + "_currency"] = (
                "label %r prints no currency -- not assumed" % printed)
    return out


STALE_BEFORE = "2023-01"


def date_key(v):
    """Normalise a CapIQ decision date to a lexicographically sortable YYYY-MM.

    THE BUG THIS EXISTS TO KILL. 'Past Rate Cases' prints decision dates as
    **MM/YYYY**, not ISO. Comparing those strings directly does two wrong things
    silently and at the same time:

        "12/2019" >= "2023-01-01"  ->  False   every case reads as STALE
        "12/2019" >  "01/2026"     ->  True    'latest' sorts by MONTH, not year

    The first blanked the staleness test for all 14 peers. The second is worse:
    it made "latest case per opco" mean "case with the highest month number", so
    a December 2019 decision beat a January 2026 one and the allowed ROE column
    was quietly built out of the wrong cases. SRE came back at 14.46% and CNP at
    13.29% -- numbers no regulator has authorised this decade, which is the only
    reason it got caught.

    Returns None for anything unparseable, which the caller treats as "no
    decision date" rather than guessing.
    """
    if v is None:
        return None
    if hasattr(v, "year") and hasattr(v, "month"):
        return "%04d-%02d" % (v.year, v.month)
    s = norm(v)
    if not s:
        return None
    m = re.match(r"^(\d{4})-(\d{1,2})", s)                    # 2026-01-15
    if m:
        return "%04d-%02d" % (int(m.group(1)), int(m.group(2)))
    m = re.match(r"^(\d{1,2})/(\d{1,2})/(\d{4})$", s)         # 01/15/2026
    if m:
        return "%04d-%02d" % (int(m.group(3)), int(m.group(1)))
    m = re.match(r"^(\d{1,2})/(\d{4})$", s)                   # 01/2026
    if m:
        return "%04d-%02d" % (int(m.group(2)), int(m.group(1)))
    m = re.match(r"^(\d{4})$", s)                             # 2026
    if m:
        return "%s-00" % m.group(1)
    return None


def read_allowed_roe(wb, extract_past, missing):
    """Rate-base-weighted authorised ROE, mirroring the ROE Bridge Tier-1 rule.

    Latest DECIDED case per (company, service type) carrying a disclosed
    auth_roe. Decisions before 2023 are STALE and excluded from the weighting --
    a 2016 allowed ROE is not a current cost of equity. Weighted by authorised
    rate base where disclosed, simple average otherwise, and the basis says which.
    Black-box settlements that disclose no ROE are counted, not silently dropped.
    """
    if "Past Rate Cases" not in wb.sheetnames:
        missing["allowed_roe_pct"] = "no 'Past Rate Cases' sheet"
        return {}
    try:
        cases = extract_past(wb["Past Rate Cases"])
    except Exception as exc:                                  # noqa: BLE001
        missing["allowed_roe_pct"] = "extract_past raised %s" % exc
        return {}
    if not cases:
        missing["allowed_roe_pct"] = "'Past Rate Cases' parsed to zero rows"
        return {}

    latest = {}
    undisclosed = 0
    undated = 0
    for c in cases:
        dd = date_key(c.get("decision_date"))
        if not dd:
            undated += 1
            continue
        key = (norm(c.get("company")), norm(c.get("service_type")))
        if c.get("auth_roe") is None:
            undisclosed += 1
            continue
        if key not in latest or dd > latest[key][0]:
            latest[key] = (dd, c)

    fresh = [(dd, c) for dd, c in latest.values() if dd >= STALE_BEFORE]
    stale = len(latest) - len(fresh)
    is_stale = False
    if not fresh:
        # Fall back to the stale set rather than returning nothing. A 2019 allowed
        # ROE is still the number the company is earning against; it is just not a
        # current cost of equity. The ROE Bridge shows these with a STALE badge and
        # excludes them from the weighting -- same treatment here, via the flag,
        # because a blank column reads as "no data" when the truth is "old data".
        if not latest:
            missing["allowed_roe_pct"] = (
                "no decided case discloses an ROE (%d black-box settlements)"
                % undisclosed)
            return {"allowed_roe_cases_undisclosed": undisclosed}
        fresh, is_stale = list(latest.values()), True

    wts = [(num(c.get("auth_rate_base")), num(c.get("auth_roe"))) for _, c in fresh]
    wts = [(w, r) for w, r in wts if r is not None]
    weighted = [(w, r) for w, r in wts if w]
    if weighted and len(weighted) >= max(1, len(wts) // 2):
        tot = sum(w for w, _ in weighted)
        roe = sum(w * r for w, r in weighted) / tot
        basis = ("rate-base weighted over %d of %d cases (%.0f%% of cases "
                 "disclose a rate base)" % (len(weighted), len(wts),
                                            100.0 * len(weighted) / len(wts)))
    else:
        roe = sum(r for _, r in wts) / len(wts)
        basis = ("simple average of %d cases -- too few disclose an "
                 "authorised rate base to weight" % len(wts))

    if is_stale:
        basis = ("STALE -- every disclosed decision predates %s. %s. Treat as "
                 "historical, not a current cost of equity." % (STALE_BEFORE, basis))
    return {
        "allowed_roe_pct": round(roe, 3),
        "allowed_roe_stale": is_stale,
        "allowed_roe_basis": basis,
        "allowed_roe_cases_used": len(wts),
        "allowed_roe_cases_stale": stale,
        "allowed_roe_cases_undisclosed": undisclosed,
        "allowed_roe_cases_undated": undated,
        "allowed_roe_latest_decision": max(dd for dd, _ in fresh),
    }


def read_ratings(wb, extract_current_ratings, missing):
    """Agency ratings off 'Current Ratings'.

    The rated ENTITY is carried with every rating. A holdco senior-unsecured and
    an opco first-mortgage rating are different numbers, and stacking them across
    a screen without saying which is which makes the column meaningless.
    """
    if "Current Ratings" not in wb.sheetnames:
        missing["credit"] = "no 'Current Ratings' sheet"
        return {}
    try:
        r = extract_current_ratings(wb["Current Ratings"])
    except Exception as exc:                                  # noqa: BLE001
        missing["credit"] = "extract_current_ratings raised %s" % exc
        return {}
    if not r:
        missing["credit"] = "'Current Ratings' parsed to nothing"
        return {}
    return {"credit": r}


SKIP_LABELS = ("source", "currency", "as of", "periods", "period ", "sort ",
               "snl financial", "data shown", "(in ", "note", "nyse", "nasdaq")


def read_planned_capex(wb, company_name, vintage_year, missing):
    r"""'Planned Capital Expenditures (As Reported)' -- captured STRUCTURALLY.

    THERE IS NO SINGLE LAYOUT. Six workbooks produced four different shapes:

      SO   entity hierarchy   Southern Company 2026..2030, then Alabama Power,
                              Georgia Power ... -- the subs SUM BACK to the first
                              block, so adding every row double counts.
      DUK  by category        Electric Generation / Transmission / Distribution /
                              Environmental, then "Total EU&I". NO parent block --
                              the first block is a COMPONENT.
      FE   by segment         Distribution / Integrated / Stand-Alone Transmission,
                              and the year labels read "2026 Forecast", not "2026".
      SRE  single period      one value per segment plus an explicit Total row,
                              no year rows at all.

    A "take the first block" rule reads Southern correctly ($78.1bn, ties to the
    sheet) and Duke WRONGLY -- it returns Electric Generation 2026-2028 as though
    it were the plan, $27.5bn against a real plan several times that. Populated,
    plausible, wrong: the same shape as `total_capacity_mw` holding the coal row.

    So a total is computed ONLY when the sheet says what the total is:
      * an explicit `Total` row, or
      * a block whose label matches the company name (the true parent).
    Otherwise every block is stored verbatim and the total is left ABSENT with
    the shape named, because choosing which components to add is a judgment call
    and inventing one silently is how the dashboard ends up quoting a number
    nobody can defend.
    """
    sheet = next((s for s in wb.sheetnames
                  if s.lower().startswith("planned capital expenditures")), None)
    if not sheet:
        missing["capex_plan"] = "no 'Planned Capital Expenditures' sheet"
        return {}
    rows = rows_of(wb, sheet)

    units, as_of, src_doc = None, None, None
    for row in rows[:26]:
        lab = norm(row[0])
        m = re.search(r"\(in (billions|millions|thousands)\)", lab, re.I)
        if m:
            units = m.group(1).lower()
        if lab.lower().startswith("as of date"):
            as_of = norm(row[1] if len(row) > 1 else None)
        if lab.lower().startswith("source document"):
            src_doc = norm(row[1] if len(row) > 1 else None)

    blocks, cur = [], None
    for i, row in enumerate(rows):
        lab = norm(row[0])
        if not lab:
            continue
        val = num(row[1] if len(row) > 1 else None)
        ym = re.match(r"^(\d{4})\b", lab)          # '2026' and '2026 Forecast'
        if ym and cur is not None:
            if val is not None:
                cur["by_year"][ym.group(1)] = val
            continue
        if i < 10 or lab.lower().startswith(SKIP_LABELS):
            continue
        cur = {"label": lab, "by_year": {}}
        if val is not None:
            cur["value"] = val                     # single-period shape (SRE)
        blocks.append(cur)

    blocks = [b for b in blocks if b["by_year"] or "value" in b]
    if not blocks:
        # Keep the vintage even when nothing parses -- otherwise this is
        # indistinguishable from "no sheet", and it is not. DTE's sheet is the
        # 2010 10-K: period labels read "2011-2013" and the values are TEXT
        # RANGES ("3.4 - 3.8"), so there is no number to take. Reporting that as
        # a missing sheet hides a sixteen-year-old vintage behind a parser
        # complaint.
        missing["capex_plan"] = (
            "'%s' is present but carries no numeric values (as-of %s, from %s). "
            "Typically a legacy sheet whose periods are spans and whose values "
            "are text ranges." % (sheet, as_of or "unstated", src_doc or "unstated"))
        return {"capex_plan": {"sheet": sheet, "as_of": as_of,
                               "source_document": src_doc,
                               "units_printed": units or "unstated",
                               "shape": "unparseable", "total_rows": [],
                               "blocks": []}}

    mult = {"billions": 1000.0, "millions": 1.0, "thousands": 0.001}.get(units)

    def total_of(b):
        return sum(b["by_year"].values()) if b["by_year"] else b.get("value")

    # A `Total` row closes the GROUP ABOVE IT, not the sheet. TXNM prints
    #   Construction expenditures 2026/2027 -> Total 10,214.8
    #   Dividends on common stock 2026/2027 -> Total 920.4
    #   Total capital requirements           -> Total 11,137.8   (incl. dividends!)
    # so "first row starting with Total" returned construction capex for two
    # years as though it were the plan. Worse, two names were wrong at the
    # HOLDCO level in a way that reads perfectly: Duke stored `Total EU&I`
    # (Electric Utilities & Infrastructure -- a SEGMENT) against a company line
    # of 58,450, and PSEG stored `Total PSE&G` (the utility SUBSIDIARY) against
    # a parent line of 14,305. Both understated, both plausible, neither wrong
    # enough to notice.
    #
    # With more than one Total on the sheet the mapping from total to scope is
    # genuinely ambiguous, so all of them are surfaced and NONE is chosen.
    totals = [b for b in blocks if re.match(r"^total\b", b["label"], re.I)]

    def block_total(b):
        return sum(b["by_year"].values()) if b["by_year"] else b.get("value")

    parent, shape = None, "components_only"
    if len(totals) == 1:
        parent, shape = totals[0], "explicit_total"
    elif len(totals) > 1:
        shape = "ambiguous_totals"
    if parent is None and shape != "ambiguous_totals":
        cname = re.sub(r"[^a-z ]", "", (company_name or "").lower())
        cname = re.sub(r"\b(the|inc|corp|corporation|company|co|group|holdings|"
                       r"energy|incorporated)\b", " ", cname)
        cname = re.sub(r"\s+", " ", cname).strip()
        for b in blocks:
            blab = re.sub(r"[^a-z ]", "", b["label"].lower())
            if cname and cname.split() and all(w in blab for w in cname.split()):
                parent, shape = b, "entity_hierarchy"
                break

    out = {
        "capex_plan": {
            "sheet": sheet,
            "as_of": as_of,
            "source_document": src_doc,
            "units_printed": units or "unstated",
            "shape": shape,
            "total_rows": [{"label": b["label"], "value": block_total(b)}
                           for b in totals],
            "blocks": blocks,
        }
    }
    if parent is None and shape == "ambiguous_totals":
        listed = " | ".join("%s = %s" % (b["label"], block_total(b))
                            for b in totals[:6])
        missing["capex_plan_total_usd_m"] = (
            "'%s' carries %d 'Total' rows and each one closes the group above it, "
            "so which is the company is ambiguous. NOT chosen. Candidates: %s. "
            "They are all in capex_plan.total_rows -- pick one deliberately."
            % (sheet, len(totals), listed))
        return out
    if parent is None:
        missing["capex_plan_total_usd_m"] = (
            "'%s' is organised by component (%s) with no consolidated line and no "
            "block matching the company name. The blocks are stored verbatim; a "
            "total needs a deliberate choice of which to add, not a guess."
            % (sheet, " / ".join(b["label"] for b in blocks[:4])))
        return out
    if mult is None:
        missing["capex_plan_total_usd_m"] = (
            "no '(in ...)' magnitude line on the sheet -- not scaled on an "
            "assumption")
        return out
    tot = total_of(parent)
    if tot is None:
        missing["capex_plan_total_usd_m"] = "parent block %r carries no values" % parent["label"]
        return out
    if as_of:
        out["capex_plan_as_of"] = as_of

    # A SINGLE-PERIOD TOTAL IS NOT A PLAN. Sempra's sheet prints one value per
    # segment for 2024 FY and a Total of 9,195 -- that is one year of capex, not
    # a multi-year programme. Writing it into a field called
    # `capex_plan_total_usd_m` would put $9.2bn beside Southern's $78.1bn
    # five-year plan in the same column and invite exactly the wrong comparison.
    out["capex_plan_total_label"] = parent["label"]

    # Vintage check applies to the single-period path too -- CEG's sheet is
    # as-of 12/31/2020 off a 2020 10-K (Exelon-era Constellation) and NWE's is
    # 12/31/2014. Both would otherwise have shipped unflagged.
    m = re.search(r"(\d{4})", as_of or "")
    if m and vintage_year and int(m.group(1)) < vintage_year - 1:
        out["capex_plan_stale"] = True
        out["capex_plan_stale_note"] = (
            "as-of %s against a %d workbook, sourced from %s -- %d years old, the "
            "plan has certainly been updated since."
            % (as_of, vintage_year, src_doc, vintage_year - int(m.group(1))))

    nyears = len(parent["by_year"])
    if nyears < 2:
        out["capex_single_period_usd_m"] = round(tot * mult, 1)
        out["capex_single_period_note"] = (
            "'%s' carries ONE period (%s), not a multi-year plan. Stored "
            "separately so it cannot be compared against a peer's plan total."
            % (sheet, as_of or "period unstated"))
        missing["capex_plan_total_usd_m"] = (
            "sheet is a single-period breakdown, not a plan -- see "
            "capex_single_period_usd_m")
        return out

    out["capex_plan_total_usd_m"] = round(tot * mult, 1)
    out["capex_plan_years"] = sorted(int(y) for y in parent["by_year"])
    out["capex_plan_basis"] = (
        "'%s' -> %s block %r, %s-%s, printed in %s and converted to $M. Component "
        "blocks are NOT added -- they sum back into this one. Compare only "
        "against peers with the same number of plan years."
        % (sheet, shape, parent["label"], min(parent["by_year"]),
           max(parent["by_year"]), units))

    return out


# ---------------------------------------------------------------------------
# precedents join
# ---------------------------------------------------------------------------

def deals_for(ticker, precedents):
    """Deals this peer appears in, matched on target/acquirer ticker.

    Derived, not entered -- precedents.json already carries a per-number source
    ledger for each of these, so re-typing them here would create a second copy
    that drifts from the audited one.
    """
    if not precedents:
        return None
    out = []
    for d in precedents:
        t = norm(d.get("target_ticker")).upper()
        a = norm(d.get("acquirer_ticker")).upper()
        if ticker not in (t, a):
            continue
        out.append({
            "deal_id": d.get("deal_id") or d.get("id"),
            "side": "acquirer" if ticker == a else "target",
            "counterparty": d.get("target") if ticker == a else d.get("acquirer"),
            "announced": d.get("announce_date") or d.get("announced"),
            "status": d.get("status"),
            "ev_usd_b": d.get("fv_usd_b"),
            "asset_class": d.get("asset_class"),
        })
    out.sort(key=lambda x: str(x.get("announced") or ""), reverse=True)
    return out


# ---------------------------------------------------------------------------
# build one peer record
# ---------------------------------------------------------------------------

# Projected down from the full coverage field set. Peers are a comp screen, not
# a coverage name: no broker research, no credit opinions, no call notes, no
# executives. Promote a name to coverage if it needs those.
KEEP_SERIES = ["periods", "revenue", "ebitda", "ebitda_margin", "net_income",
               "eps_diluted", "net_debt", "total_debt", "total_equity", "capex",
               "dividends_ps", "roe", "total_assets", "net_ppe"]
KEEP_SCALAR = ["name", "total_capacity_mw", "total_capacity_operating_mw",
               "coal_capacity_mw", "gas_capacity_mw", "oil_capacity_mw",
               "hydro_capacity_mw", "nuclear_capacity_mw", "wind_capacity_mw",
               "solar_capacity_mw", "renewable_capacity_mw",
               "electric_customers", "gas_customers", "fte_employees",
               "rate_base_best", "rate_base_best_label", "rate_base_best_company",
               "rate_base_consolidated", "rate_base_consolidated_label"]


def build_record(path, ticker, exchange, mods):
    vintage = vintage_of(path)
    vintage_year = int(vintage[:4]) if vintage else None
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    try:
        missing = {}
        ext = mods["eaf"].Extractor(wb)
        ext.run()
        v = ext.values

        rec = collections.OrderedDict()
        rec["ticker"] = ticker
        rec["tier"] = "peer"
        rec["name"] = v.get("name")
        rec["exchange"] = exchange
        rec["asset_class"] = ASSET_CLASS.get(ticker)
        if rec["asset_class"] is None:
            missing["asset_class"] = ("ticker not in the ASSET_CLASS map -- add it "
                                      "deliberately rather than letting it default")

        cur = (ext.currency or {}).get("financials") or []
        rec["currency"] = cur[0] if len(set(cur)) == 1 and cur else None
        if rec["currency"] is None:
            missing["currency"] = ("workbook does not print a single reported "
                                   "currency code (%r)" % cur)
        expected = NON_USD.get(ticker, "USD")
        if rec["currency"] and rec["currency"] != expected:
            rec["_currency_flag"] = ("workbook reports %s, expected %s -- check "
                                     "before comparing absolute figures"
                                     % (rec["currency"], expected))

        for f in KEEP_SERIES + KEEP_SCALAR:
            if f in v and f != "name":
                rec[f] = v[f]
        for f in KEEP_SERIES + KEEP_SCALAR:
            if f not in rec and f in ext.missing:
                missing[f] = ext.missing[f]

        rec.update(read_capitalisation(wb, missing))
        rec.update(read_valuation(wb, missing))
        rec.update(read_allowed_roe(wb, mods["ea"].extract_past, missing))
        rec.update(read_ratings(wb, mods["ea"].extract_current_ratings, missing))
        rec.update(read_planned_capex(wb, v.get("name"), vintage_year, missing))

        # ── EV / rate base ────────────────────────────────────────────────────
        # THREE separate traps here, all of which produce a populated, plausible-
        # looking, wrong number -- the same shape as `total_capacity_mw` holding
        # the coal row for eleven of eighteen tickers.
        #
        #  1. UNITS. `tev_000` is THOUSANDS of currency. `rate_base_best` is
        #     dollars (extract_aqn_financials multiplies the sheet's $M by 1e6).
        #     Dividing them directly is wrong by 1,000x.
        #  2. NUMERATOR SCOPE. `rate_base_best` is the LARGEST SINGLE pending
        #     opco, not the company. Summing the opcos is the better proxy.
        #  3. COVERAGE. Even the sum only counts opcos with a PENDING case, so it
        #     understates any company that is not currently in rate cases
        #     everywhere. This is a proxy, and the field says so.
        opcos = v.get("rate_base_opcos") or {}
        rb_sum_m = sum(o.get("rate_base_m") or 0 for o in opcos.values()) or None
        if rb_sum_m:
            rec["rate_base_pending_sum_usd_m"] = round(rb_sum_m, 1)
            rec["rate_base_pending_opco_count"] = len(opcos)
        nppe = v.get("net_ppe")
        nppe_latest = None
        if isinstance(nppe, list):
            nppe_latest = next((num(x) for x in reversed(nppe)
                                if num(x) is not None), None)
        cov = (rb_sum_m * 1000.0 / nppe_latest) if (rb_sum_m and nppe_latest) else None
        if cov is not None:
            rec["rate_base_coverage_vs_net_ppe"] = round(cov, 3)

        # ── EV / rate base is NOT computed here. Deliberately. ────────────────
        # Three rounds of trying to salvage it, each one narrowing the gate:
        #
        #   round 1  no gate           FE 5.44x, XEL 4.92x
        #   round 2  0.3x-6.0x band    the band was wider than the error
        #   round 3  >=40% coverage    let WEC 2.89x, SWX 2.45x, NJR 2.34x through
        #
        # Round 3 is the instructive one. The denominator only counts opcos with a
        # PENDING case, so a multiple computed at 53% coverage is overstated by
        # roughly 1/0.53. WEC's stored 2.89x against a real rate base near $29bn is
        # about 2.0x -- wrong by 45%, and wrong in a way that still LOOKS like a
        # utility multiple, which is the dangerous kind. DTE showed the opposite
        # failure: 100% "coverage" means the pending figure equals net PP&E, which
        # is not what a rate base does, so 1.69x is understated.
        #
        # The error is proportional to a coverage number that varies 3%-100% across
        # the book, so there is no threshold that makes this column safe. The
        # ingredients are kept -- `rate_base_pending_sum_usd_m` and
        # `rate_base_coverage_vs_net_ppe` -- and the multiple waits for a
        # consolidated rate base from RRA or company disclosure. A wrong number is
        # worse than an empty field.
        if rb_sum_m and cov is not None:
            missing["ev_rate_base"] = (
                "NOT COMPUTED. Pending-case rate base of %.0f $M across %d opco(s) "
                "is %.0f%% of net PP&E, so any multiple off it is wrong by roughly "
                "1/%.2f. Needs a consolidated rate base (RRA or company "
                "disclosure)." % (rb_sum_m, len(opcos), 100 * cov, cov))
        else:
            missing["ev_rate_base"] = (
                "NOT COMPUTED -- no consolidated rate base available. Pending-case "
                "rate base is a partial denominator; see rate_base_pending_sum_usd_m.")

        # Dividend yield. Deliberately the LATEST REPORTED FY, not a forward
        # yield: an 'NA' in the current period is CapIQ stating there is no
        # dividend, and reaching back past it to the last number that exists is
        # exactly what put XIFR on the dashboard at a 22.82% yield after XPLR
        # suspended its distribution outright.
        dps, px = rec.get("dividends_ps"), rec.get("close_price")
        if isinstance(dps, list) and dps and px:
            last = dps[-1]
            if last is None or (isinstance(last, str) and last.strip().upper() in NULLISH):
                missing["div_yield_pct"] = (
                    "latest FY dividend prints NA -- the company pays none. NOT "
                    "backfilled from an earlier year.")
                rec["dividend_status"] = "no dividend in the latest reported FY"
            else:
                lv = num(last)
                if lv is not None and rec.get("close_price_currency") == rec.get("currency"):
                    rec["div_yield_pct"] = round(100.0 * lv / px, 3)
                    rec["div_yield_basis"] = (
                        "latest reported FY DPS / latest close, both %s. This is a "
                        "TRAILING yield, not a forward one." % rec.get("currency"))
                elif lv is not None:
                    missing["div_yield_pct"] = (
                        "price is %s while DPS is %s -- not computed"
                        % (rec.get("close_price_currency"), rec.get("currency")))
        else:
            missing.setdefault("div_yield_pct", "no dividend series or no close price")

        rec["source_workbook"] = os.path.basename(path)
        rec["workbook_vintage"] = vintage
        rec["_missing"] = collections.OrderedDict(sorted(missing.items()))
        return rec
    finally:
        wb.close()


# ---------------------------------------------------------------------------
# diff + write
# ---------------------------------------------------------------------------

def populated(rec):
    out = set()
    for k, v in rec.items():
        if k.startswith("_"):
            continue
        if v is None or v == [] or v == {} or v == "":
            continue
        out.add(k)
    return out


def diff_and_report(before, after):
    losses = []
    for t, old in before.items():
        new = after.get(t)
        if new is None:
            losses.append((t, "<entire record>"))
            continue
        gone = populated(old) - populated(new)
        losses.extend((t, f) for f in sorted(gone))
    return losses


def atomic_write(path, doc):
    tmp = path + ".tmp"
    with open(tmp, "w", encoding="utf-8") as fh:
        json.dump(doc, fh, ensure_ascii=False, indent=2)
    with open(tmp, encoding="utf-8") as fh:
        json.load(fh)                       # re-parse before the swap
    os.replace(tmp, path)


# ---------------------------------------------------------------------------

def main(argv=None):
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--peer-reports", required=True)
    ap.add_argument("--out", required=True)
    ap.add_argument("--scripts-dir", default=os.path.dirname(os.path.abspath(__file__)))
    ap.add_argument("--precedents")
    ap.add_argument("--dry-run", action="store_true")
    ap.add_argument("--allow-loss", action="store_true")
    a = ap.parse_args(argv)

    check_guards(a.peer_reports, a.out)

    mods = {
        "eaf": load_module(os.path.join(a.scripts_dir, "extract_aqn_financials.py"),
                           "eaf"),
        "ea": load_module(os.path.join(a.scripts_dir, "extract_all.py"), "ea"),
    }

    precedents = None
    if a.precedents and os.path.isfile(a.precedents):
        doc = json.load(open(a.precedents, encoding="utf-8"))
        precedents = doc.get("deals", doc) if isinstance(doc, dict) else doc
        print("precedents: %d deals loaded for the recent_deals join" % len(precedents))
    elif a.precedents:
        print("precedents: %r not found -- recent_deals will be absent" % a.precedents)

    before = {}
    if os.path.isfile(a.out):
        doc = json.load(open(a.out, encoding="utf-8"),
                        object_pairs_hook=collections.OrderedDict)
        before = doc.get("peers", doc)
    after = collections.OrderedDict(
        (t, collections.OrderedDict(r)) for t, r in before.items())

    files = sorted(f for f in os.listdir(a.peer_reports)
                   if f.endswith(".xlsx") and not f.startswith("~$"))
    if not files:
        die("no .xlsx workbooks in %r" % a.peer_reports)

    collisions = []
    for fn in files:
        t, _ = ticker_of(fn)
        if t in COVERAGE_UNIVERSE:
            collisions.append((t, fn))
    if collisions:
        for t, fn in collisions:
            print("  COLLISION  %-6s %s" % (t, fn))
        die("%d workbook(s) belong to the 24-name COVERAGE universe. Peers must "
            "be names OUTSIDE coverage -- --peer-reports is almost certainly "
            "pointed at data\\reports\\." % len(collisions))

    print("\n%-6s %-9s %7s %7s  %s" % ("ticker", "class", "found", "missing", "workbook"))
    print("-" * 92)
    for fn in files:
        t, exch = ticker_of(fn)
        if not t:
            print("  [skip] %s -- no ticker in filename" % fn[:60])
            continue
        try:
            rec = build_record(os.path.join(a.peer_reports, fn), t, exch, mods)
        except Exception as exc:                              # noqa: BLE001
            print("  [FAIL] %-6s %s: %s" % (t, fn[:40], exc))
            continue
        d = deals_for(t, precedents)
        if d is not None:
            rec["recent_deals"] = d
        after[t] = rec
        print("%-6s %-9s %7d %7d  %s"
              % (t, rec.get("asset_class") or "?", len(populated(rec)),
                 len(rec["_missing"]), fn[:44]))

    losses = diff_and_report(before, after)
    if losses:
        print("\n%d FIELD LOSS(ES):" % len(losses))
        for t, f in losses[:40]:
            print("   %-6s %s" % (t, f))
        if not a.allow_loss:
            die("a pre-existing peer would lose a populated field. Re-run with "
                "--allow-loss once you have read the list above. Every merge "
                "reports success either way -- this diff is the only thing that "
                "does not.")
    else:
        print("\ndiff: no pre-existing peer loses a populated field")

    doc = collections.OrderedDict()
    doc["schema"] = "peer_universe/1.0"
    doc["tier"] = ("Tier 2 lightweight comps -- names OUTSIDE the 24-name coverage "
                   "universe. No broker research, credit opinions, call notes or "
                   "executives by design.")
    doc["peers"] = collections.OrderedDict(sorted(after.items()))
    doc["count"] = len(after)

    print("\n%d peers (%d new this run)" % (len(after), len(after) - len(before)))
    if a.dry_run:
        print("DRY RUN -- nothing written")
        return 0
    atomic_write(a.out, doc)
    print("wrote %s (atomic replace, re-parsed before swap)" % a.out)
    print("\nNothing else was touched. capiq_export.json and executives_export.json "
          "were never opened.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
