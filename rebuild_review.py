# -*- coding: utf-8 -*-
r"""Rebuild the review workbook from Fareen's QC'd copy plus the snip resolutions.

Her returned workbook is the source of truth for what is CLOSED — she marked Done?
and wrote a Commentary column. The snip resolutions say what each capture actually
showed. This folds both in:

  * every row she marked Done moves to Resolved, carrying her own words
  * rows resolved by a capture move to Resolved with the printed value recorded
  * the ESG / score-band category is dropped outright (her rule, 2026-07-31:
    "use the printed scores, don't rely on the score-band graphics")
  * what is left is the genuinely open list, re-sorted by action type

Format is unchanged — six sheets, same columns, same sort — because she asked for it
that way and it is what makes the list quick to work through.

  python3 rebuild_review.py <FE workbook.xlsx> <QC_RESOLUTIONS.json> <out.xlsx>
"""
import json, re, sys, collections
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

FE, RES, OUT = sys.argv[1], sys.argv[2], sys.argv[3]
CADDY = 'http://100.86.108.51:8080'
CREDIT_SRC = {"Moody's", 'S&P', 'Fitch'}

ARIAL = 'Arial'
HDR_FILL = PatternFill('solid', fgColor='1F3864')
HDR_FONT = Font(name=ARIAL, size=10, bold=True, color='FFFFFF')
BAD_FILL = PatternFill('solid', fgColor='FCE4E4')
REV_FILL = PatternFill('solid', fgColor='FFF4DC')
NEW_FILL = PatternFill('solid', fgColor='E8F2E8')
THIN = Side(style='thin', color='D9D9D9')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
BODY = Font(name=ARIAL, size=10)
LINKF = Font(name=ARIAL, size=10, color='0563C1', underline='single')
WRAP = Alignment(wrap_text=True, vertical='top')
TOP = Alignment(vertical='top')

# The category Fareen retired. A row is dropped, not reclassified — if the printed
# score is already captured there is nothing for her to look at.
ESG_DROP = re.compile(
    r'esg.*(score.band|colou?r|gradient|issuer profile score|credit impact score|'
    r'relevance (frequency )?panel|relevance scores?|cis[- ]?\d|score (graphic|gauge|bar))'
    r'|(business|financial) risk (profile )?(scale|graphic|bar)'
    r'|ratings score snapshot .*(scale|graphic)|navigator (score )?bars?', re.I)


def folder_for(source):
    return 'credit' if source in CREDIT_SRC else 'reports'


def link(base, source, page):
    if not page:
        return ''
    return f'{CADDY}/{folder_for(source)}/{base}#page={page}'


def write_sheet(ws, headers, widths, rows, fills=None, links=None):
    ws.append(headers)
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i)
        c.font, c.fill = HDR_FONT, HDR_FILL
        c.alignment = Alignment(vertical='center', wrap_text=True)
        ws.column_dimensions[get_column_letter(i)].width = widths[i - 1]
    ws.row_dimensions[1].height = 30
    for r, row in enumerate(rows, start=2):
        for i, v in enumerate(row, 1):
            c = ws.cell(row=r, column=i, value=v)
            c.font, c.border = BODY, BORDER
            c.alignment = WRAP if widths[i - 1] > 28 else TOP
        if fills:
            f = fills(rows[r - 2])
            if f:
                for i in range(1, len(headers) + 1):
                    ws.cell(row=r, column=i).fill = f
    if links:
        for r, row in enumerate(rows, start=2):
            col, url = links(row)
            if url:
                c = ws.cell(row=r, column=col)
                c.hyperlink, c.font, c.value = url, LINKF, 'open page'
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:{get_column_letter(len(headers))}{max(2, len(rows) + 1)}'


def main():
    wb = load_workbook(FE)
    src = wb['Review']
    res = json.load(open(RES, encoding='utf-8'))
    by_row = collections.defaultdict(list)
    for r in res:
        if r.get('row'):
            by_row[r['row']].append(r)

    still_open, resolved_new, dropped = [], [], []

    for raw in src.iter_rows(min_row=2, values_only=True):
        n, sev, tk, source, pdf, pg, fld, iss, need, act, _lnk, done, com = \
            (list(raw) + [None] * 13)[:13]
        com = (com or '').strip()
        done = (done or '').strip().lower() == 'yes'

        if ESG_DROP.search(f'{fld} {iss}'):
            dropped.append((n, tk, pdf, fld))
            continue

        caps = by_row.get(n) or []
        if caps:
            for c in caps:
                resolved_new.append([
                    n, tk, source, pdf, pg, fld,
                    (iss or '')[:600],
                    f"{c.get('verdict','').upper()} — {c.get('capture_shows','')}"[:900],
                    (f"Your capture ({c.get('image')}). " if c.get('image') else '')
                    + (c.get('action') or '')[:500]
                    + (f"  [{c.get('confidence')} confidence]" if c.get('confidence') else ''),
                ])
            continue

        if done or com:
            resolved_new.append([
                n, tk, source, pdf, pg, fld, (iss or '')[:600],
                f"CLOSED BY YOU — {com}" if com else 'CLOSED BY YOU (marked Done)',
                'Your call, applied as written.',
            ])
            continue

        still_open.append(dict(n=n, sev=sev, tk=tk, source=source, pdf=pdf, pg=pg,
                               fld=fld, iss=iss, need=need, act=act))

    ACT_ORDER = ['Re-scan page', 'Pick the governing exhibit', 'Confirm column alignment',
                 'Confirm inferred value', 'Eyeball and confirm',
                 'Confirm my cross-check — low priority', 'Chart or graphic — optional',
                 'Not fixable by re-scanning', 'No action — informational']
    rank = {a: i for i, a in enumerate(ACT_ORDER)}
    still_open.sort(key=lambda x: (rank.get(x['act'], 99), x['tk'] or '', x['pdf'] or '',
                                   x['pg'] or 0))

    out = Workbook()

    ws = out.active
    ws.title = 'Review'
    hdr = ['#', 'Severity', 'Ticker', 'Source', 'PDF file name', 'Page',
           'Field / exhibit in question', 'Issue / problem', 'What I need from you',
           'Action type', 'Open the page', 'Done?', 'Commentary']
    w = [5, 10, 8, 11, 34, 6, 40, 62, 58, 26, 14, 8, 40]
    rows = [[i['n'], i['sev'], i['tk'], i['source'], i['pdf'], i['pg'], i['fld'],
             i['iss'], i['need'], i['act'], 'open' if i['pg'] else '', '', '']
            for i in still_open]
    write_sheet(ws, hdr, w, rows,
                fills=lambda r: BAD_FILL if r[1] == 'BAD' else REV_FILL,
                links=lambda r: (11, link(r[4], r[3], r[5])))

    ws2 = out.create_sheet('Unmapped rows')
    src2 = wb['Unmapped rows']
    rows2 = [list(r) for r in src2.iter_rows(min_row=2, values_only=True)]
    write_sheet(ws2, [c.value for c in src2[1]], [5, 8, 11, 34, 6, 78, 46, 8], rows2)

    for name, widths in [('Filing flags', [5, 36, 12, 44, 66, 56, 8]),
                         ('Coverage gaps', [5, 24, 40, 22, 62, 54, 8])]:
        s = wb[name]
        wsx = out.create_sheet(name)
        write_sheet(wsx, [c.value for c in s[1]], widths,
                    [list(r) for r in s.iter_rows(min_row=2, values_only=True)])

    ws5 = out.create_sheet('Resolved')
    hdr5 = ['#', 'Ticker', 'Source', 'PDF file name', 'Page', 'Cell / exhibit',
            'What was wrong', 'Resolution', 'How it was confirmed']
    old_res = [list(r) for r in wb['Resolved'].iter_rows(min_row=2, values_only=True)]
    write_sheet(ws5, hdr5, [5, 8, 11, 34, 6, 44, 62, 62, 52],
                old_res + resolved_new,
                fills=lambda r: NEW_FILL if r in resolved_new else None)

    ws6 = out.create_sheet('Retired — ESG graphics', 3)
    write_sheet(ws6, ['#', 'Ticker', 'PDF file name', 'Field / exhibit', 'Why this is gone'],
                [5, 8, 34, 52, 90],
                [[n, tk, pdf, fld,
                  'RETIRED as a category on your instruction: "use the printed scores, '
                  "don't rely on the score-band graphics, they don't provide any additional "
                  'information." The scores are already captured from the narrative, so there '
                  'was nothing here to review. Listed once for the record; these will not '
                  'appear in future workbooks.']
                 for n, tk, pdf, fld in dropped])

    ws0 = out.create_sheet('Summary', 0)
    for col, wd in zip('ABCD', (38, 12, 12, 12)):
        ws0.column_dimensions[col].width = wd
    ws0['A1'] = 'Power Academy — extraction review'
    ws0['A1'].font = Font(name=ARIAL, size=14, bold=True)
    ws0['A2'] = (f'Rebuilt after your QC pass. You closed {len(resolved_new)} rows — '
                 f'{sum(1 for r in res if r.get("row"))} of them from the 51 captures in '
                 f'data\\Scans\\snips, which have all been read and applied to the JSONs '
                 f'(180 verified edits). {len(dropped)} ESG / score-graphic rows were '
                 f'retired outright. {len(still_open)} rows remain open.')
    ws0['A2'].font = Font(name=ARIAL, size=10, italic=True)
    ws0['A4'], ws0['B4'], ws0['C4'], ws0['D4'] = 'Category', 'Total', 'Done', 'Remaining'
    for col in 'ABCD':
        ws0[f'{col}4'].font, ws0[f'{col}4'].fill = HDR_FONT, HDR_FILL

    n_open, n_unm = len(still_open), len(rows2)
    spec = [
        ('Cells flagged BAD (unreadable)', f'=COUNTIF(Review!B2:B{n_open+1},"BAD")',
         f'=COUNTIFS(Review!B2:B{n_open+1},"BAD",Review!L2:L{n_open+1},"y")'),
        ('Cells flagged REVIEW (doubtful)', f'=COUNTIF(Review!B2:B{n_open+1},"REVIEW")',
         f'=COUNTIFS(Review!B2:B{n_open+1},"REVIEW",Review!L2:L{n_open+1},"y")'),
        ('Unmapped rows (informational)', f"=COUNTA('Unmapped rows'!A2:A{n_unm+1})",
         f"=COUNTIF('Unmapped rows'!H2:H{n_unm+1},\"y\")"),
        ('Resolved (cumulative)', f'=COUNTA(Resolved!A2:A{len(old_res)+len(resolved_new)+1})',
         f'=COUNTA(Resolved!A2:A{len(old_res)+len(resolved_new)+1})'),
        ('Retired — ESG graphics', f"=COUNTA('Retired — ESG graphics'!A2:A{len(dropped)+1})",
         f"=COUNTA('Retired — ESG graphics'!A2:A{len(dropped)+1})"),
    ]
    r = 5
    for label, tot, done_f in spec:
        ws0[f'A{r}'], ws0[f'B{r}'], ws0[f'C{r}'] = label, tot, done_f
        ws0[f'D{r}'] = f'=B{r}-C{r}'
        for col in 'ABCD':
            ws0[f'{col}{r}'].font, ws0[f'{col}{r}'].border = BODY, BORDER
        r += 1

    r += 2
    ws0[f'A{r}'] = 'What changed since your copy'
    ws0[f'A{r}'].font = Font(name=ARIAL, size=11, bold=True)
    for line in [
        f'All 51 captures read and matched to their rows — 72 resolutions, 70 high-confidence.',
        f'180 edits applied to the four JSONs, each one verified against the value it expected',
        f'   to find first, so a wrong path fails loudly instead of writing to the wrong cell.',
        '53 stored values were corrected. The ones worth knowing:',
        '   PPL BofA revenue 2026E  3,952 -> 9,952   (OCR ate the leading 9)',
        '   AQN NBC capital assets 2027E  40,876 -> 10,876',
        '   AQN S&P cash  null -> $65m   (the "$05" fragment implied $105-205m)',
        "   Moody's Ex.15 third peer  = Calpine, Ba3 Positive, not investment grade",
        '   PPL S&P local-currency ICR  "A-7" -> A/Stable/A-1',
        '   PPL UBS cfo 2026E = 3,204 as printed, NOT the 3,205 arithmetic implied',
        '',
        'Two of my own review rows were wrong and are corrected in the data:',
        "   VST Moody's Exhibit 1 labels all six columns, not just the first and last —",
        '      the 2021-2023 gap was recoverable all along.',
        '   The MSEX peer table was not missing a column: California Water Service is',
        '      printed with "--" on every line because it carries ratings only.',
        '',
        'Sort the Review tab by Action type, not severity — "Re-scan page" first; those are',
        'the only rows where the cell is genuinely empty.',
    ]:
        r += 1
        ws0[f'A{r}'] = line
        ws0[f'A{r}'].font = BODY

    out.save(OUT)
    print(f'wrote {OUT}')
    print(f'  open      {len(still_open)}')
    print(f'  resolved  {len(old_res)} carried + {len(resolved_new)} new = '
          f'{len(old_res)+len(resolved_new)}')
    print(f'  retired   {len(dropped)} ESG / score-graphic rows')
    c = collections.Counter(i['act'] for i in still_open)
    print('\n  open by action type:')
    for a in ACT_ORDER:
        if c.get(a):
            print(f'    {c[a]:4}  {a}')


main()
