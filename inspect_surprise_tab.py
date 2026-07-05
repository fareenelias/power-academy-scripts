"""
inspect_surprise_tab.py  — fixed
- Matches exact 'Surprise' sheet name first
- Avoids ReadOnlyWorksheet.dimensions (not available in read_only mode)
Usage: python inspect_surprise_tab.py ETR
"""
import openpyxl, os, sys

REPORTS_DIR = r"E:\PowerAcademy\data\reports"
ticker = sys.argv[1].upper() if len(sys.argv) > 1 else 'ETR'

files = [f for f in os.listdir(REPORTS_DIR) if ticker in f.upper() and f.endswith('.xlsx')]
if not files:
    print(f"No file for {ticker}"); sys.exit(1)

path = os.path.join(REPORTS_DIR, files[0])
print(f"File: {files[0]}\n")

wb = openpyxl.load_workbook(path, data_only=True, read_only=True)

# Exact match first, then substring
sheet_name = next((n for n in wb.sheetnames if n == 'Surprise'), None)
if not sheet_name:
    sheet_name = next((n for n in wb.sheetnames if n.lower() == 'surprise'), None)
if not sheet_name:
    print("No exact 'Surprise' sheet found.")
    print("Sheets with 'surprise':", [n for n in wb.sheetnames if 'surprise' in n.lower()])
    sys.exit(1)

ws = wb[sheet_name]
print(f"Sheet: {sheet_name}\n")
print("First 25 rows x 20 cols:")
for i, row in enumerate(ws.iter_rows(max_row=25, values_only=True)):
    vals = [str(v)[:22] if v is not None else '' for v in list(row)[:20]]
    if any(vals):
        print(f"  R{i+1:02d}: {vals}")