"""
diag_ma_sheets.py  —  Paste output back to Claude.
Shows first 40 non-empty rows of M&A sheets for AEE (large utility)
and YORW (small utility) to confirm column layout before building extractor.
"""
import os, glob
import openpyxl

REPORTS_DIR = r"E:\PowerAcademy\data\reports"
TARGET_SHEETS = ['Transactions Summary', 'Detailed M&A History']

files = sorted(glob.glob(os.path.join(REPORTS_DIR, '*.xlsx')))

# Pick one large utility (AEE or NEE) and one small (YORW)
samples = {}
for f in files:
    base = os.path.basename(f)
    for ticker in ['AEE', 'NEE', 'D', 'ETR']:
        if ticker in base and ticker not in samples:
            samples[ticker] = f
    for ticker in ['YORW', 'MSEX', 'AWR']:
        if ticker in base and ticker not in samples:
            samples[ticker] = f
    if len(samples) >= 4:
        break

for ticker, fpath in samples.items():
    print(f"\n{'='*70}")
    print(f"FILE: {os.path.basename(fpath)}")
    print('='*70)
    wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)

    for sheet_name in TARGET_SHEETS:
        if sheet_name not in wb.sheetnames:
            print(f"\n  ['{sheet_name}' NOT FOUND]")
            continue
        ws = wb[sheet_name]
        print(f"\n--- Sheet: '{sheet_name}' ---")
        count = 0
        for row in ws.iter_rows(values_only=True):
            vals = [str(v)[:70] if v is not None else '' for v in row]
            if any(v.strip() for v in vals):
                print(f"  R{count+1}: {vals[:10]}")  # show first 10 cols
                count += 1
            if count >= 40:
                print("  ...")
                break

    wb.close()