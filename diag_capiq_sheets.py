"""
diag_capiq_sheets.py
Lists all sheets in the CapIQ Excel reports and finds rate case data.
Run: python E:\PowerAcademy\scripts\diag_capiq_sheets.py
"""
import os, glob
try:
    import openpyxl
except ImportError:
    os.system("pip install openpyxl --break-system-packages -q")
    import openpyxl

REPORTS_DIR = r'E:\PowerAcademy\data\reports'
files = sorted(glob.glob(os.path.join(REPORTS_DIR, '*.xlsx')))
print(f"Found {len(files)} Excel files in {REPORTS_DIR}\n")

# Step 1: List all sheet names for the first 3 files
print("=== SHEET NAMES (first 3 files) ===")
for fpath in files[:3]:
    wb = openpyxl.load_workbook(fpath, data_only=True, read_only=True)
    print(f"\n{os.path.basename(fpath)}:")
    for i, sname in enumerate(wb.sheetnames):
        print(f"  [{i:2d}] {sname}")
    wb.close()

# Step 2: For NEE file, dump first 10 rows of EVERY sheet to find rate cases
nee_file = next((f for f in files if 'NextEra' in f or 'NEE' in f), files[0])
print(f"\n\n=== ALL SHEETS — {os.path.basename(nee_file)} (first rows each) ===")
wb = openpyxl.load_workbook(nee_file, data_only=True, read_only=True)
for sname in wb.sheetnames:
    ws = wb[sname]
    rows = list(ws.iter_rows(max_row=8, values_only=True))
    non_empty = [(i, [str(v)[:35] for v in r if v is not None])
                 for i, r in enumerate(rows, 1) if any(v is not None for v in r)]
    if non_empty:
        print(f"\n  Sheet: '{sname}'")
        for rn, vals in non_empty[:4]:
            print(f"    R{rn}: {vals}")
wb.close()

# Step 3: Search ALL sheets in ALL files for rate case keywords
print("\n\n=== RATE CASE KEYWORD SEARCH (all files, all sheets) ===")
RC_KEYWORDS = ['rate case', 'pending case', 'past case', 'authorized roe',
               'requested roe', 'equity ratio', 'equity layer', 'docket',
               'test year', 'rate base approved', 'allowed roe', 'equity thickness']
hits = []
for fpath in files:
    try:
        wb = openpyxl.load_workbook(fpath, data_only=True, read_only=True)
        for sname in wb.sheetnames:
            ws = wb[sname]
            rows = list(ws.iter_rows(max_row=200, values_only=True))
            for i, row in enumerate(rows[:200], 1):
                row_str = ' '.join(str(v or '').lower() for v in row)
                if any(kw in row_str for kw in RC_KEYWORDS):
                    hits.append((os.path.basename(fpath), sname, i, row))
                    break  # one hit per sheet is enough to identify it
        wb.close()
    except Exception as e:
        print(f"  Error {os.path.basename(fpath)}: {e}")

if hits:
    print(f"Found {len(hits)} sheets with rate case data:")
    for fname, sname, rn, row in hits:
        vals = [str(v)[:40] for v in row if v is not None]
        print(f"  {fname[:50]} | '{sname}' | R{rn}: {vals[:5]}")

    # Deep dive: show 30 rows from first hit
    best = hits[0]
    bfname = next(f for f in files if os.path.basename(f) == best[0])
    print(f"\n\n=== DEEP DIVE: {best[0]} / '{best[1]}' ===")
    wb = openpyxl.load_workbook(bfname, data_only=True, read_only=True)
    ws = wb[best[1]]
    rows = list(ws.iter_rows(max_row=50, values_only=True))
    for i, row in enumerate(rows, 1):
        vals = [str(v)[:35] for v in row if v is not None]
        if vals:
            print(f"  R{i:3d}: {vals}")
    wb.close()
else:
    print("No rate case keywords found. Dumping sheet names of all files:")
    for fpath in files:
        try:
            wb = openpyxl.load_workbook(fpath, data_only=True, read_only=True)
            print(f"  {os.path.basename(fpath)}: {wb.sheetnames}")
            wb.close()
        except Exception as e:
            print(f"  {os.path.basename(fpath)}: ERROR {e}")