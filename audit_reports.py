"""
audit_reports.py — Full raw audit of every CapIQ Excel report

Purpose: extract_all.py only pulls fields we've explicitly coded for.
This script instead opens EVERY report and dumps EVERY row of EVERY sheet,
so we can see the complete data surface — what we're already using and
what's sitting there unused — before deciding what to wire into the
real extractor.

Usage:
    cd E:\\PowerAcademy
    python scripts\\audit_reports.py

Outputs (written to E:\\PowerAcademy\\data\\audit\\):
    - audit_<ticker>.json      one per company: every sheet, every non-empty row
    - _toc.csv                 master table of contents: every sheet in every
                                 workbook, with row/col counts and a header preview
    - _labels.json             for each sheet name (e.g. "Credit Ratios (x)"),
                                 the union of column-A/B labels seen across all
                                 companies that have that sheet — fastest way to
                                 see "what fields exist in this tab across our coverage"

Run this FIRST, review the outputs together, THEN decide what to add to
extract_all.py. This script does not modify capiq_export.json or
executives_export.json.
"""

import json
import re
import sys
import traceback
from pathlib import Path
from collections import defaultdict

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)

# ---------------------------------------------------------------------------
# CONFIG
# ---------------------------------------------------------------------------

REPORTS_DIR = Path(r"E:\PowerAcademy\data\reports")
OUTPUT_DIR = Path(r"E:\PowerAcademy\data\audit")

# CapIQ export filenames follow a consistent pattern:
#   {CompanyName}{EXCHANGE}{TICKER}_Report_{date}.xlsx
# e.g. "Evergy,Inc.NASDAQGSEVRG_Report_06-28-2026.xlsx" -> EVRG
# Longer/more-specific exchange codes MUST come before shorter ones in this
# list (NASDAQGS before NASDAQ) or the regex grabs the wrong split point —
# this is the same bug noted previously where EVRG extracted as "GSEVRG".
EXCHANGE_PREFIXES = ["NASDAQGS", "NASDAQGM", "NASDAQCM", "NASDAQ", "NYSEAMERICAN", "NYSE"]

TICKER_PATTERN = re.compile(
    r"(?:" + "|".join(EXCHANGE_PREFIXES) + r")([A-Z]{1,6})_Report", re.IGNORECASE
)

EXCEL_EXTENSIONS = {".xlsx", ".xlsm", ".xls"}


def guess_ticker(filename: str) -> str:
    stem = Path(filename).stem
    m = TICKER_PATTERN.search(stem)
    if m:
        return m.group(1).upper()
    # fall back to filename stem so nothing gets dropped/overwritten silently —
    # if you see one of these in the output, the filename didn't match the
    # expected {Company}{EXCHANGE}{TICKER}_Report_{date} pattern and is worth
    # a manual look
    return stem[:40]


def cell_str(value):
    """Normalize a cell value to something JSON-safe and readable."""
    if value is None:
        return None
    if hasattr(value, "isoformat"):  # datetime/date
        return value.isoformat()
    return value


def dump_workbook(path: Path):
    """Return {sheet_name: [{row: int, values: [...]}]} for every non-empty row."""
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    sheets_out = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows_out = []
        for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            values = [cell_str(v) for v in row]
            # skip fully empty rows to cut noise, but keep the row index
            # so gaps are still visible if you need exact positions later
            if all(v is None or (isinstance(v, str) and v.strip() == "") for v in values):
                continue
            # trim trailing Nones from the row for readability
            while values and values[-1] is None:
                values.pop()
            rows_out.append({"row": row_idx, "values": values})
        sheets_out[sheet_name] = rows_out
    wb.close()
    return sheets_out


def main():
    if not REPORTS_DIR.exists():
        print(f"ERROR: reports directory not found: {REPORTS_DIR}")
        sys.exit(1)

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    files = [
        p for p in REPORTS_DIR.iterdir()
        if p.suffix.lower() in EXCEL_EXTENSIONS and not p.name.startswith("~$")
    ]
    files.sort()

    if not files:
        print(f"No Excel files found in {REPORTS_DIR}")
        sys.exit(1)

    print(f"Found {len(files)} report(s) in {REPORTS_DIR}\n")

    toc_rows = []                       # for _toc.csv
    labels_by_sheet = defaultdict(lambda: defaultdict(set))  # sheet_name -> label -> {tickers}
    ticker_seen = defaultdict(list)     # ticker -> [filenames] to flag collisions

    for i, path in enumerate(files, start=1):
        ticker = guess_ticker(path.name)
        ticker_seen[ticker].append(path.name)
        print(f"[{i}/{len(files)}] {path.name}  ->  {ticker}")

        try:
            sheets = dump_workbook(path)
        except Exception as e:
            print(f"    !! FAILED to read: {e}")
            traceback.print_exc()
            continue

        out_path = OUTPUT_DIR / f"audit_{ticker}.json"
        payload = {
            "source_file": path.name,
            "ticker_guess": ticker,
            "sheet_count": len(sheets),
            "sheets": sheets,
        }
        with open(out_path, "w", encoding="utf-8") as f:
            json.dump(payload, f, indent=2, ensure_ascii=False)

        for sheet_name, rows in sheets.items():
            header_preview = rows[0]["values"][:6] if rows else []
            toc_rows.append({
                "ticker": ticker,
                "source_file": path.name,
                "sheet_name": sheet_name,
                "row_count": len(rows),
                "header_preview": header_preview,
            })
            # collect column A / column B as candidate "labels" — matches the
            # vertical label layout most CapIQ tabs use (per known extraction notes)
            for r in rows:
                vals = r["values"]
                if not vals:
                    continue
                label = vals[0] if vals[0] else (vals[1] if len(vals) > 1 else None)
                if isinstance(label, str) and 0 < len(label.strip()) < 80:
                    labels_by_sheet[sheet_name][label.strip()].add(ticker)

        print(f"    wrote {out_path.name}  ({len(sheets)} sheets)")

    # ---- flag any ticker collisions (two files mapped to the same ticker) ----
    collisions = {t: fns for t, fns in ticker_seen.items() if len(fns) > 1}
    if collisions:
        print("\n!! TICKER COLLISIONS — same ticker guessed for multiple files (review these):")
        for t, fns in collisions.items():
            print(f"   {t}: {fns}")

    # ---- write _toc.csv ----
    import csv
    toc_path = OUTPUT_DIR / "_toc.csv"
    with open(toc_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["ticker", "source_file", "sheet_name", "row_count", "header_preview"])
        for row in toc_rows:
            writer.writerow([
                row["ticker"], row["source_file"], row["sheet_name"],
                row["row_count"], json.dumps(row["header_preview"], ensure_ascii=False),
            ])

    # ---- write _labels.json ----
    labels_out = {
        sheet_name: {
            label: sorted(tickers) for label, tickers in sorted(label_map.items())
        }
        for sheet_name, label_map in sorted(labels_by_sheet.items())
    }
    labels_path = OUTPUT_DIR / "_labels.json"
    with open(labels_path, "w", encoding="utf-8") as f:
        json.dump(labels_out, f, indent=2, ensure_ascii=False)

    print(f"\nDone. {len(files)} workbook(s) audited.")
    print(f"  Per-company dumps: {OUTPUT_DIR}\\audit_<ticker>.json")
    print(f"  Table of contents: {toc_path}")
    print(f"  Label inventory:   {labels_path}")
    print("\nNext: open _toc.csv first for a quick skim of every sheet across every "
          "company, then _labels.json to see which fields recur within a given sheet "
          "type. We'll use both to decide what extract_all.py should pick up next.")


if __name__ == "__main__":
    main()