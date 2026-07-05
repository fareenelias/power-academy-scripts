"""
inspect_diversified_utility.py
Shows the structure of the 'Diversified Utility Highlights' sheet in ETR's CapIQ Excel.
This is where per-opco rate base, ROE, equity thickness live for diversified utilities.
"""
import openpyxl, os

REPORTS_DIR = r"E:\PowerAcademy\data\reports"
files = [f for f in os.listdir(REPORTS_DIR) if 'ETR' in f.upper() and f.endswith('.xlsx')]
path  = os.path.join(REPORTS_DIR, files[0])
print(f"File: {files[0]}\n")

wb = openpyxl.load_workbook(path, data_only=True, read_only=True)

target = next((n for n in wb.sheetnames if 'Diversified Utility Highlights' in n), None)
if not target:
    print("Sheet not found. Available sheets with 'Utility':")
    for n in wb.sheetnames:
        if 'utility' in n.lower() or 'Utility' in n:
            print(f"  {n}")
else:
    ws = wb[target]
    print(f"Sheet: {target}\nFirst 40 rows x 15 cols:\n")
    for i, row in enumerate(ws.iter_rows(max_row=40, values_only=True)):
        vals = [str(v)[:20] if v is not None else '' for v in list(row)[:15]]
        if any(vals):
            print(f"  R{i+1:02d}: {vals}")