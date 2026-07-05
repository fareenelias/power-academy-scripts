"""
diag_ma_detail.py  —  Paste output back to Claude.
Checks full column width of Detailed M&A History data rows,
and prints Advisers + Advisory & Underwriting Summary sheets.
"""
import os, glob
import openpyxl

REPORTS_DIR = r"E:\PowerAcademy\data\reports"

# Use D (Dominion - many deals) + ETR (also many deals) + one small company
WANT = ['D', 'ETR', 'YORW']
found = {}
for f in sorted(glob.glob(os.path.join(REPORTS_DIR, '*.xlsx'))):
    base = os.path.basename(f)
    for t in WANT:
        # match ticker at end of exchange prefix
        import re
        m = re.search(r'(?:NASDAQGS|NASDAQGM|NASDAQ|NYSE)' + t + r'_Report', base)
        if m and t not in found:
            found[t] = f

for ticker, fpath in found.items():
    print(f"\n{'='*70}")
    print(f"TICKER: {ticker}  FILE: {os.path.basename(fpath)}")
    print('='*70)
    wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
    sheets = wb.sheetnames

    # 1. Full columns of first 5 data rows from Detailed M&A History
    if 'Detailed M&A History' in sheets:
        ws = wb['Detailed M&A History']
        print("\n--- Detailed M&A History: FULL ROW for first 5 SPTRD rows ---")
        count = 0
        for row in ws.iter_rows(values_only=True):
            col0 = str(row[0]).strip() if row[0] else ''
            if col0.startswith('SPTRD'):
                # Show all non-None columns
                vals = [(i, str(v)[:50]) for i, v in enumerate(row) if v is not None]
                print(f"  cols present: {[i for i,v in vals]}")
                for i, v in vals:
                    print(f"    [{i}] {v}")
                print()
                count += 1
                if count >= 3:
                    break

    # 2. Advisers sheet
    for sheet_name in ['Advisers', 'Advisory & Underwriting Summary']:
        if sheet_name in sheets:
            ws = wb[sheet_name]
            print(f"\n--- Sheet: '{sheet_name}' (first 50 non-empty rows) ---")
            count = 0
            for row in ws.iter_rows(values_only=True):
                vals = [str(v)[:60] if v is not None else '' for v in row]
                if any(v.strip() for v in vals):
                    print(f"  R{count+1}: {vals[:8]}")
                    count += 1
                if count >= 50:
                    print("  ...")
                    break
        else:
            print(f"\n  ['{sheet_name}' NOT in sheets]")

    wb.close()