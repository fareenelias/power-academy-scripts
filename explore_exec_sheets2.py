"""
explore_exec_sheets2.py  —  Fixed version. Paste output back to Claude.
"""
import os, glob
import openpyxl

REPORTS_DIR = r"E:\PowerAcademy\data\reports"
TARGET_SHEETS = ['People', 'Corporate Governance', 'Compensation']

files = sorted(glob.glob(os.path.join(REPORTS_DIR, "*.xlsx")))
# Use Ameren as sample (electric), then a water utility
samples = files[:1]
water = [f for f in files if any(t in os.path.basename(f).upper()
         for t in ['MSEX','AWK','CWT','AWR'])]
if water:
    samples.append(water[0])

for fpath in samples:
    print(f"\n{'='*70}")
    print(f"FILE: {os.path.basename(fpath)}")
    print('='*70)
    wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)

    for sheet_name in TARGET_SHEETS:
        if sheet_name not in wb.sheetnames:
            print(f"\n  [Sheet '{sheet_name}' NOT FOUND]")
            continue
        ws = wb[sheet_name]
        print(f"\n--- Sheet: '{sheet_name}' ---")
        row_num = 0
        for row in ws.iter_rows(values_only=True):
            row_num += 1
            vals = [str(v)[:80] if v is not None else '' for v in row]
            if any(v.strip() for v in vals):  # skip fully empty rows
                print(f"  R{row_num}: {vals}")
            if row_num >= 60:  # enough to see the structure
                print(f"  ... (stopped at row {row_num})")
                break

    wb.close()