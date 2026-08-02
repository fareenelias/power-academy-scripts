# -*- coding: utf-8 -*-
r"""Rebuild the generation-capacity fields in capiq_export.json from every CapIQ
workbook's "Current Capacity Summary" sheet.

Written 2026-07-31 after Fareen asked why AQN showed no hydro: "AQN sold its
renewables, but it should have hydro assets for generation capacity." She was right,
and chasing it turned up two problems, the second much larger than the first.

1. THERE IS NO HYDRO FIELD AT ALL. The schema carried solar / wind / nuclear / gas
   and nothing else, so hydro was dropped for EVERY ticker - AQN's 121.85 MW, but
   also NEE's 1,300.9 MW of pumped storage and VST's 108.5 MW. Coal and oil were
   dropped the same way.

2. `total_capacity_mw` WAS THE COAL ROW. The sheet lists "Total Coal" first and the
   fleet "Total" last; whatever populated this field took the first row whose label
   started with "Total". So NEE stored 222.75 MW against a real fleet of 131,046 MW,
   VST stored 9,224.13 against 49,232.63, and so on. Every cross-name capacity
   comparison in the dashboard has been wrong. AQN was the odd one out only because
   it was filled later, correctly, from the actual Total row.

Column convention: the pre-existing solar / wind / nuclear values match column H,
"Total Capacity (MW)", which INCLUDES under-construction and planned. That is kept so
nothing silently shifts, but operating nameplate (column C) is stored alongside as
`*_operating_mw`, because the gap is not small - NEE is 83,670 MW operating against
131,046 MW including planned, and a reader comparing "fleet size" deserves to know
which one they are looking at.

  python3 extract_capacity.py --reports <dir> --capiq <capiq_export.json> [--dry-run]

Merge-only: no field outside the capacity set is touched.
"""
import argparse, json, os, re, sys, collections

try:
    import openpyxl
except ImportError:
    sys.exit('openpyxl required: pip install openpyxl --break-system-packages')

SHEET = 'Current Capacity Summary'

# printed row label -> field stem. Matched on the normalised label, longest first,
# so "Total Natural Gas" cannot be caught by a looser "gas" rule.
ROWS = [
    ('total coal',                    'coal'),
    ('total natural gas',             'gas'),
    ('oil & other petroleum products','oil'),
    ('uranium',                       'nuclear'),
    ('total hydro',                   'hydro'),
    ('total renewable',               'renewable'),
    ('other non-renewable',           'other_nonrenewable'),
    ('wind',                          'wind'),
    ('solar',                         'solar'),
    ('biomass',                       'biomass'),
]
COL_NAMEPLATE, COL_TOTAL = 3, 8   # C = Operating Nameplate, H = Total Capacity


def norm(x):
    return re.sub(r'\s+', ' ', str(x or '')).strip().lower()


def num(v):
    if v is None:
        return None
    s = str(v).strip().replace(',', '')
    if s.upper() in ('NA', 'N/A', '', '-', '--'):
        return None
    try:
        return float(s)
    except ValueError:
        return None


def ticker_of(fname):
    base = re.sub(r'_Report_.*$', '', os.path.splitext(os.path.basename(fname))[0])
    for pre in ('NASDAQGS', 'NASDAQGM', 'NASDAQCM', 'NASDAQ', 'NYSEARCA', 'NYSEAMERICAN',
                'NYSEMKT', 'NYSE', 'TSXV', 'TSX', 'AMEX', 'OTCPK'):
        if pre in base:
            t = base.split(pre)[-1]
            if t:
                return t
    return None


def read_capacity(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        if SHEET not in wb.sheetnames:
            return None, 'no "%s" sheet' % SHEET
        ws = wb[SHEET]
        rows = [(norm(r[0]), r) for r in ws.iter_rows(min_row=1, max_row=80,
                                                      max_col=COL_TOTAL, values_only=True)]
    finally:
        wb.close()

    out, seen = {}, set()
    # sub-rows are indented under their parent, e.g. "        Solar". Both the parent
    # ("Total Renewable") and the child ("Solar") are wanted, so match on the
    # normalised label and take the FIRST occurrence of each.
    for label, r in rows:
        if not label:
            continue
        for want, stem in ROWS:
            if stem in seen:
                continue
            if label == want or label.startswith(want + ':'):
                tot = num(r[COL_TOTAL - 1])
                nam = num(r[COL_NAMEPLATE - 1])
                if tot is not None:
                    out[stem + '_capacity_mw'] = tot
                if nam is not None:
                    out[stem + '_capacity_operating_mw'] = nam
                seen.add(stem)
                break

    # The FLEET total is the LAST bare "Total" row, not the first "Total ..." row.
    # This is the bug: taking the first match lands on "Total Coal".
    fleet = [r for label, r in rows if label == 'total']
    if fleet:
        r = fleet[-1]
        tot, nam = num(r[COL_TOTAL - 1]), num(r[COL_NAMEPLATE - 1])
        if tot is not None:
            out['total_capacity_mw'] = tot
        if nam is not None:
            out['total_capacity_operating_mw'] = nam
    return out, None


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--reports', required=True)
    ap.add_argument('--capiq', required=True)
    ap.add_argument('--dry-run', action='store_true')
    a = ap.parse_args()

    data = json.load(open(a.capiq, encoding='utf-8'),
                     object_pairs_hook=collections.OrderedDict)
    companies = data.get('companies', data)

    files = sorted(f for f in os.listdir(a.reports)
                   if f.endswith('.xlsx') and not f.startswith('~$'))
    changed = collections.Counter()
    report = []

    for fn in files:
        t = ticker_of(fn)
        if not t or t not in companies:
            print('  [skip] %s (ticker %r not in capiq)' % (fn[:52], t))
            continue
        cap, err = read_capacity(os.path.join(a.reports, fn))
        if err:
            print('  [skip] %-6s %s' % (t, err))
            continue
        before = companies[t].get('total_capacity_mw')
        for k, v in cap.items():
            if companies[t].get(k) != v:
                changed[k] += 1
            companies[t][k] = v
        report.append((t, before, cap.get('total_capacity_mw'),
                       cap.get('hydro_capacity_mw'), cap.get('coal_capacity_mw')))

    print('\n%-6s %14s %14s %10s %10s   %s' %
          ('ticker', 'total WAS', 'total NOW', 'hydro', 'coal', 'verdict'))
    print('-' * 88)
    for t, was, now, hyd, coal in sorted(report):
        if was is not None and coal is not None and abs(was - coal) < 0.01 and (now or 0) != was:
            verdict = 'WAS THE COAL ROW'
        elif was == now:
            verdict = 'unchanged'
        elif was is None:
            verdict = 'newly filled'
        else:
            verdict = 'corrected'
        print('%-6s %14s %14s %10s %10s   %s' %
              (t, was if was is not None else '-', now if now is not None else '-',
               hyd if hyd is not None else '-', coal if coal is not None else '-', verdict))

    print('\nfields written:', dict(changed))
    if a.dry_run:
        print('\nDRY RUN - nothing written')
        return
    tmp = a.capiq + '.tmp'
    json.dump(data, open(tmp, 'w', encoding='utf-8'), ensure_ascii=False, indent=2)
    json.load(open(tmp, encoding='utf-8'))
    os.replace(tmp, a.capiq)
    print('\nwrote %s (atomic replace, re-parsed before swap)' % a.capiq)


main()
