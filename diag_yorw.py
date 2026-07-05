"""
diag_yorw.py  —  Paste output back to Claude.
"""
import json, os
import openpyxl

CAPIQ_FILE  = r"E:\PowerAcademy\data\capiq_export.json"
REPORTS_DIR = r"E:\PowerAcademy\data\reports"

# ── 1. What's in capiq_export.json for YORW? ─────────────────────────────────
with open(CAPIQ_FILE, encoding='utf-8') as f:
    capiq = json.load(f)
companies = capiq.get('companies', capiq)

print("=== YORW in capiq_export.json ===")
if 'YORW' not in companies:
    print("  NOT FOUND in companies dict")
else:
    co = companies['YORW']
    for k, v in co.items():
        if isinstance(v, list):
            print(f"  {k}: {v}")
        elif k not in ('pending_rate_cases', 'past_rate_cases', 'earnings_surprise'):
            print(f"  {k}: {v}")
    print(f"  earnings_surprise: {len(co.get('earnings_surprise', []))} entries")
    print(f"  pending_rate_cases: {len(co.get('pending_rate_cases', []))} entries")
    print(f"  past_rate_cases: {len(co.get('past_rate_cases', []))} entries")

# ── 2. What sheets does the YORW Excel file have? ────────────────────────────
print("\n=== YORW Excel sheets ===")
yorw_file = None
for f in os.listdir(REPORTS_DIR):
    if 'YORW' in f.upper() and f.endswith('.xlsx'):
        yorw_file = os.path.join(REPORTS_DIR, f)
        print(f"  File: {f}")
        break

if yorw_file:
    wb = openpyxl.load_workbook(yorw_file, read_only=True, data_only=True)
    print(f"  Sheets: {wb.sheetnames}")

    # Check Income Statement for EPS rows
    for sheet_name in ['Income Statement', 'Consensus', 'Mean Estimates and Actuals Summ',
                       'Detailed Estimates', 'Surprise']:
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            print(f"\n  --- {sheet_name} (first 30 non-empty rows) ---")
            count = 0
            for row in ws.iter_rows(values_only=True):
                vals = [str(v)[:60] if v is not None else '' for v in row]
                if any(v.strip() for v in vals):
                    print(f"    {vals[:8]}")
                    count += 1
                    if count >= 30:
                        print("    ...")
                        break
    wb.close()
else:
    print("  YORW Excel file not found")