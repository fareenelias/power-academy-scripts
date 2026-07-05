"""
diag_ratecase_excel.py  v2
Finds CapIQ Excel reports and dumps rate case table structure.
Also prints extract_capiq_reports.py rate case section.
Run: python E:\PowerAcademy\scripts\diag_ratecase_excel.py
"""
import os, glob, sys

# 1. Print the extractor's rate case section
EXTRACTOR = r'E:\PowerAcademy\scripts\extract_capiq_reports.py'
if os.path.exists(EXTRACTOR):
    print("=== extract_capiq_reports.py (rate case related lines) ===")
    with open(EXTRACTOR, encoding='utf-8', errors='ignore') as f:
        lines = f.readlines()
    rc_kw = ['rate_case','pending','past','authorized_roe','rate_base',
             'equity','docket','decision_date','requested_roe','authorized_rev']
    rc_lines = [(i+1, l.rstrip()) for i, l in enumerate(lines)
                if any(kw in l.lower() for kw in rc_kw)]
    for lineno, line in rc_lines[:60]:
        print(f"  L{lineno:4d}: {line}")
    print(f"\n  ({len(rc_lines)} matching lines total)")
else:
    print(f"Extractor NOT found at {EXTRACTOR}")

# 2. Search for Excel files
print("\n\n=== Searching for CapIQ Excel files ===")
DIRS = [
    r'E:\PowerAcademy',
    r'E:\PowerAcademy\data',
    r'E:\PowerAcademy\data\capiq_reports',
    r'E:\PowerAcademy\reports',
    os.path.expanduser('~\\Downloads'),
    os.path.expanduser('~\\Desktop'),
    os.path.expanduser('~\\Documents'),
]
all_xl = []
for d in DIRS:
    if os.path.exists(d):
        for ext in ['*.xlsx','*.xls','*.xlsm']:
            found = glob.glob(os.path.join(d,'**',ext), recursive=True)
            if found:
                print(f"  {d}: {len(found)} {ext}")
            all_xl.extend(found)
all_xl = list(set(all_xl))
print(f"Total: {len(all_xl)} Excel files")
for f in sorted(all_xl, key=os.path.getsize, reverse=True)[:20]:
    print(f"  [{os.path.getsize(f)//1024:>5}KB] {f}")

if not all_xl:
    print("\nNo Excel files found. Check the path to your CapIQ reports.")
    sys.exit(0)

# 3. Inspect for rate case data
print("\n\n=== Inspecting for rate case data ===")
try:
    import openpyxl
except ImportError:
    os.system("pip install openpyxl --break-system-packages -q")
    import openpyxl

all_xl.sort(key=os.path.getsize, reverse=True)
found_rc = 0
for fpath in all_xl[:15]:
    try:
        wb = openpyxl.load_workbook(fpath, data_only=True, read_only=True)
        for sname in wb.sheetnames:
            ws = wb[sname]
            rows = list(ws.iter_rows(max_row=120, values_only=True))
            rc_idx = [i for i, r in enumerate(rows)
                      if any(kw in str(r).lower() for kw in
                             ['rate case','authorized','docket','requested roe',
                              'equity','test year','pending','decision'])]
            if rc_idx:
                print(f"\n{'='*55}")
                print(f"FILE: {os.path.basename(fpath)}")
                print(f"SHEET: '{sname}' — hit at rows {[i+1 for i in rc_idx[:5]]}")
                first = rc_idx[0]
                for i in range(max(0,first-1), min(len(rows),first+20)):
                    r = rows[i]
                    vals = [str(v)[:28] for v in r if v is not None]
                    if vals: print(f"  R{i+1:3d}: {vals}")
                found_rc += 1
                if found_rc >= 3: break
        if found_rc >= 3: break
    except Exception as e:
        pass

if found_rc == 0:
    print("\nNo rate case sheets found. Sheet names in top files:")
    for fpath in all_xl[:5]:
        try:
            wb = openpyxl.load_workbook(fpath, data_only=True, read_only=True)
            print(f"\n  {os.path.basename(fpath)}: {wb.sheetnames}")
            for sname in wb.sheetnames[:2]:
                ws = wb[sname]
                for i, r in enumerate(ws.iter_rows(max_row=5, values_only=True)):
                    vals = [str(v)[:25] for v in r if v is not None]
                    if vals: print(f"    {sname} R{i+1}: {vals}")
        except Exception as e:
            print(f"  {os.path.basename(fpath)}: {e}")