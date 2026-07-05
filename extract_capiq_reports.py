"""
extract_capiq_reports.py
Reads all CapIQ Excel reports in data/reports/ and extracts structured data
into capiq_export.json.

Run: python E:\PowerAcademy\scripts\extract_capiq_reports.py

Requires: pip install openpyxl --break-system-packages
"""

import os, re, glob, json
from datetime import datetime

try:
    import openpyxl
except ImportError:
    import subprocess, sys
    subprocess.check_call([sys.executable, '-m', 'pip', 'install', 'openpyxl', '--break-system-packages', '-q'])
    import openpyxl

# ── Config ──────────────────────────────────────────────────────────────────
REPORTS_DIR = r'E:\PowerAcademy\data\reports'
CAPIQ_JSON  = r'E:\PowerAcademy\data\capiq_export.json'

# ── Helpers ─────────────────────────────────────────────────────────────────

def ticker_from_filename(fname):
    """Extract ticker from filenames like 'NextEraEnergy,Inc.NYSENEE_Report_06-28-2026.xlsx'"""
    base = os.path.splitext(os.path.basename(fname))[0]
    # Remove '_Report_...' suffix
    base = re.sub(r'_Report_.*$', '', base)
    # Remove exchange prefixes (longest first)
    for prefix in ['NASDAQGS', 'NASDAQGM', 'NASDAQ', 'NYSEARCA', 'NYSE', 'TSX']:
        if prefix in base:
            ticker = base.split(prefix)[-1]
            if ticker:
                return ticker
    return None


def safe(v):
    """Return float, string date, or None. Never crash."""
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.strftime('%m/%Y')
    s = str(v).strip()
    if s.lower() in ('na', 'n/a', 'none', '', '-', '--', '—'):
        return None
    # Try float
    try:
        return float(s.replace(',', ''))
    except ValueError:
        pass
    # Try parsing date-like strings  
    for fmt in ('%Y-%m-%d %H:%M:%S', '%Y-%m-%d', '%m/%d/%Y', '%m/%Y'):
        try:
            dt = datetime.strptime(s.split(' ')[0] if ' ' in s else s, fmt.split(' ')[0])
            return dt.strftime('%m/%Y')
        except ValueError:
            continue
    return s  # return as string


def safe_str(v):
    if v is None: return None
    s = str(v).strip()
    return None if s.lower() in ('na','n/a','none','','-') else s


def safe_float(v):
    r = safe(v)
    if isinstance(r, float): return r
    if isinstance(r, str):
        try: return float(r.replace(',','').replace('%','').replace('$',''))
        except: pass
    return None


def safe_date(v):
    r = safe(v)
    if r is None: return None
    if isinstance(r, str) and re.match(r'\d{2}/\d{4}', r): return r
    if isinstance(r, str):
        # Try common date formats
        for fmt in ('%Y-%m-%d', '%m/%d/%Y', '%m/%Y', '%Y'):
            try:
                return datetime.strptime(r[:10] if len(r) > 10 else r, fmt).strftime('%m/%Y')
            except: continue
    return None


def ws_rows(ws, max_row=500):
    return list(ws.iter_rows(max_row=max_row, values_only=True))


# ── Pending Rate Cases (sheet: 'Pending Rate Cases') ────────────────────────
#
# Header at row 8 (index 7):
#  0:State | 1:Company | 2:Ticker | 3:Docket | 4:Service | 5:CaseType |
#  6:FilingDate | 7:RateIncrease($M) | 8:RevChange% | 9:ROCCR% | 10:ROE% |
#  11:Equity% | 12:TestYear | 13:RateBase($M) | 14:ValuationMethod | 15:ExpectedDecision

PENDING_COLS = {
    'state':            0,
    'company':          1,
    'docket':           3,
    'service_type':     4,
    'case_type':        5,
    'filing_date':      6,
    'req_revenue':      7,
    'req_rev_pct':      8,
    'req_roccr':        9,
    'req_roe':          10,
    'req_equity':       11,
    'test_year':        12,
    'req_rate_base':    13,
    'rate_base_method': 14,
    'decision_date':    15,
}

def extract_pending(ws):
    rows = ws_rows(ws)
    # Find header row
    hdr_idx = None
    for i, row in enumerate(rows):
        vals = [str(v or '').lower() for v in row]
        if 'state' in vals and 'docket' in vals and 'expected decision date' in ' '.join(vals):
            hdr_idx = i
            break
    if hdr_idx is None:
        return []

    results = []
    for row in rows[hdr_idx + 1:]:
        if not any(v for v in row):
            continue
        state = safe_str(row[0] if len(row) > 0 else None)
        if not state:
            continue

        c = {}
        for field, col in PENDING_COLS.items():
            if col >= len(row):
                c[field] = None
                continue
            v = row[col]
            if field in ('filing_date', 'test_year', 'decision_date'):
                c[field] = safe_date(v)
            elif field in ('req_revenue', 'req_rev_pct', 'req_roccr', 'req_roe', 'req_equity', 'req_rate_base'):
                c[field] = safe_float(v)
            elif field in ('state', 'company', 'docket', 'service_type', 'case_type', 'rate_base_method'):
                c[field] = safe_str(v)
            else:
                c[field] = safe(v)

        if c.get('company') or c.get('docket'):
            results.append(c)

    return results


# ── Past Rate Cases (sheet: 'Past Rate Cases') ──────────────────────────────
#
# Two header rows at R5+R6 (indices 4+5). Data from R7 (index 6).
#
# Combined column mapping:
#  0:State | 1:Company | 2:Ticker | 3:Docket | 4:Service | 5:CaseType
#  REQUESTED (under "Increase Requested"):
#  6:FilingDate | 7:ReqRevenue($M) | 8:ReqROCCR% | 9:ReqROE% | 10:ReqEquity% | 11:ReqRateBase($M)
#  AUTHORIZED (under "Increase Authorized"):
#  12:DecisionDate | 13:DecisionType | 14:AuthRevenue($M) | 15:PhaseIn | 16:Interim |
#  17:AuthROCCR% | 18:AuthROE% | 19:AuthEquity% | 20:TestYear | 21:AuthRateBase($M)
#  22:RateBaseValuationMethod | 23:Duration(months)

PAST_COLS = {
    'state':            0,
    'company':          1,
    'docket':           3,
    'service_type':     4,
    'case_type':        5,
    'filing_date':      6,
    'req_revenue':      7,
    'req_roccr':        8,
    'req_roe':          9,
    'req_equity':       10,
    'req_rate_base':    11,
    'decision_date':    12,
    'decision_type':    13,
    'auth_revenue':     14,
    'phase_in':         15,
    'interim':          16,
    'auth_roccr':       17,
    'auth_roe':         18,
    'auth_equity':      19,
    'test_year':        20,
    'auth_rate_base':   21,
    'rate_base_method': 22,
    'duration_months':  23,
}

def extract_past(ws):
    rows = ws_rows(ws)
    # Find data start — look for first row after the double header where col 0 is a state name
    data_start = None
    for i, row in enumerate(rows):
        vals = [str(v or '').lower() for v in row[:3]]
        # R5 has top-level header 'State', R6 has 'Date'
        # Data row has an actual state name
        if i >= 5 and vals[0] and vals[0] not in ('state', 'date', '') and len(vals[0]) > 1:
            data_start = i
            break
    if data_start is None:
        return []

    results = []
    for row in rows[data_start:]:
        if not any(v for v in row):
            continue
        state = safe_str(row[0] if len(row) > 0 else None)
        if not state or state.lower() in ('state', 'date'):
            continue

        c = {}
        for field, col in PAST_COLS.items():
            if col >= len(row):
                c[field] = None
                continue
            v = row[col]
            if field in ('filing_date', 'decision_date', 'test_year'):
                c[field] = safe_date(v)
            elif field in ('req_revenue', 'req_roccr', 'req_roe', 'req_equity', 'req_rate_base',
                           'auth_revenue', 'auth_roccr', 'auth_roe', 'auth_equity', 'auth_rate_base',
                           'duration_months'):
                c[field] = safe_float(v)
            elif field in ('state', 'company', 'docket', 'service_type', 'case_type',
                           'decision_type', 'phase_in', 'interim', 'rate_base_method'):
                c[field] = safe_str(v)
            else:
                c[field] = safe(v)

        if c.get('company') or c.get('docket'):
            results.append(c)

    return results


# ── Main ────────────────────────────────────────────────────────────────────

def main():
    # Load existing capiq_export.json
    if os.path.exists(CAPIQ_JSON):
        with open(CAPIQ_JSON, 'r', encoding='utf-8') as f:
            capiq = json.load(f)
        companies = capiq.get('companies', capiq)
        print(f"Loaded {CAPIQ_JSON}: {len(companies)} companies")
    else:
        companies = {}
        capiq = {'companies': companies}
        print(f"Starting fresh capiq_export.json")

    files = sorted(glob.glob(os.path.join(REPORTS_DIR, '*.xlsx')))
    # Exclude temp files (Excel lock files start with ~$)
    files = [f for f in files if not os.path.basename(f).startswith('~$')]
    print(f"\nProcessing {len(files)} Excel files from {REPORTS_DIR}\n")

    stats = {'pending': 0, 'past': 0, 'skipped': 0}

    for fpath in files:
        ticker = ticker_from_filename(fpath)
        if not ticker:
            print(f"  [SKIP] Could not extract ticker from: {os.path.basename(fpath)}")
            stats['skipped'] += 1
            continue

        print(f"  {ticker:6s} — {os.path.basename(fpath)[:55]}", end='')

        try:
            wb = openpyxl.load_workbook(fpath, data_only=True, read_only=True)
        except Exception as e:
            print(f"\n         ERROR opening file: {e}")
            stats['skipped'] += 1
            continue

        pending, past = [], []

        if 'Pending Rate Cases' in wb.sheetnames:
            ws = wb['Pending Rate Cases']
            pending = extract_pending(ws)

        if 'Past Rate Cases' in wb.sheetnames:
            ws = wb['Past Rate Cases']
            past = extract_past(ws)

        wb.close()

        # Ensure ticker exists in companies dict
        if ticker not in companies:
            companies[ticker] = {'ticker': ticker}

        companies[ticker]['pending_rate_cases'] = pending
        companies[ticker]['past_rate_cases']    = past

        print(f" | pending={len(pending):2d}  past={len(past):2d}")
        stats['pending'] += len(pending)
        stats['past']    += len(past)

    # Write back
    if 'companies' not in capiq:
        capiq = {'companies': companies}
    else:
        capiq['companies'] = companies

    with open(CAPIQ_JSON, 'w', encoding='utf-8') as f:
        json.dump(capiq, f, indent=2, ensure_ascii=False)

    print(f"\n{'='*55}")
    print(f"Done. Total pending: {stats['pending']}, past: {stats['past']}")
    print(f"Wrote: {CAPIQ_JSON}")

    # Spot-check: print sample past case to verify field mapping
    sample_ticker = next((t for t in ['NEE','D','CMS','PPL'] if t in companies), None)
    if sample_ticker:
        sample_past = companies[sample_ticker].get('past_rate_cases', [])
        if sample_past:
            print(f"\n--- Sample past case ({sample_ticker}) ---")
            c = sample_past[0]
            for k, v in c.items():
                if v is not None:
                    print(f"  {k:20s}: {v}")


if __name__ == '__main__':
    main()