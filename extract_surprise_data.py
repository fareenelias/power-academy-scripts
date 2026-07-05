"""
extract_surprise_data.py
Reads the 'Surprise' sheet from each CapIQ Excel file and writes
earnings_surprise into capiq_export.json per ticker.

Structure confirmed from ETR inspection:
  R14: ['Per Share Level', 'FQ1 2025 - Mar 2025', 'FQ2 2025 - Jun 2025', ...]
  R16: ['EPS Normalized', <surprise_pct>, ...]  <- parent metric row
  R17: ['    %',           <surprise_pct>, ...]  <- sub: same
  R18: ['    Difference($)',<surprise_d>, ...]   <- sub: $ beat/miss
  R19: ['    Actual ($)',  <actual>, ...]         <- sub: actual EPS
  R20: ['    Estimate($)', <estimate>, ...]       <- sub: consensus estimate
  R21: ['    Announced Date', <date>, ...]
  FY column included as a separate period.
"""
import openpyxl, os, json, re

REPORTS_DIR = r"E:\PowerAcademy\data\reports"
CAPIQ_FILE  = r"E:\PowerAcademy\data\capiq_export.json"

def safe_float(v):
    if v is None: return None
    try:
        n = float(str(v).replace('%','').replace('$','').replace(',','').strip())
        return None if (n == 0 and str(v).strip() == '0') else n
    except: return None

def safe_str(v):
    if v is None: return None
    s = str(v).strip()
    return s if s and s != 'NA' else None

def parse_period(label):
    """Returns (period_key, period_type, quarter, year)"""
    s = str(label).strip()
    # e.g. 'FQ1 2025 - Mar 2025' or 'FY 2025 - Dec 2025'
    if s.startswith('FQ'):
        m = re.match(r'FQ(\d) (\d{4})', s)
        if m: return s, 'quarterly', int(m.group(1)), int(m.group(2))
    if s.startswith('FY'):
        m = re.match(r'FY (\d{4})', s)
        if m: return s, 'annual', None, int(m.group(1))
    return s, 'unknown', None, None

def extract_surprise_sheet(ws):
    rows = list(ws.iter_rows(max_row=60, values_only=True))
    
    # Find period header row
    header_idx = None
    for i, row in enumerate(rows):
        if row[0] and 'Per Share Level' in str(row[0]):
            header_idx = i; break
    if header_idx is None: return []

    # Period labels from cols 1 onward (skip empty)
    raw_periods = [(j+1, v) for j, v in enumerate(rows[header_idx][1:]) if v and str(v).strip()]
    if not raw_periods: return []
    col_indices = [c for c, _ in raw_periods]
    period_labels = [v for _, v in raw_periods]

    # Parse through rows after header looking for 'EPS Normalized' block
    surprise_pct = {}; surprise_d = {}; actual = {}; estimate = {}; announced = {}
    in_eps = False

    for row in rows[header_idx + 1:]:
        label = str(row[0]).strip() if row[0] else ''
        stripped = label.lstrip()

        if any(k in stripped for k in ('EPS Normalized', 'EPS (GAAP)', 'EPS Diluted')) and not label.startswith(' '):
            in_eps = True
            for ci, pl in zip(col_indices, period_labels):
                surprise_pct[pl] = safe_float(row[ci]) if ci < len(row) else None
            continue

        if not in_eps: continue

        if 'Difference' in stripped and '$' in stripped:
            for ci, pl in zip(col_indices, period_labels):
                surprise_d[pl] = safe_float(row[ci]) if ci < len(row) else None
        elif 'Actual' in stripped and '$' in stripped:
            for ci, pl in zip(col_indices, period_labels):
                actual[pl] = safe_float(row[ci]) if ci < len(row) else None
        elif 'Estimate' in stripped and '$' in stripped:
            for ci, pl in zip(col_indices, period_labels):
                estimate[pl] = safe_float(row[ci]) if ci < len(row) else None
        elif 'Announced' in stripped:
            for ci, pl in zip(col_indices, period_labels):
                announced[pl] = safe_str(row[ci]) if ci < len(row) else None
        elif stripped and not stripped.startswith(('%','Difference','Actual','Estimate',
                                                   'Announced','Guidance','Accounting',
                                                   'Guidance')):
            break  # new metric block started

    # Build sorted list of periods
    results = []
    for pl in period_labels:
        key, ptype, quarter, year = parse_period(pl)
        results.append({
            'period':           key,
            'period_type':      ptype,
            'quarter':          quarter,
            'year':             year,
            'eps_actual':       actual.get(pl),
            'eps_estimate':     estimate.get(pl),
            'eps_surprise_d':   surprise_d.get(pl),
            'eps_surprise_pct': surprise_pct.get(pl),
            'announced_date':   announced.get(pl),
        })
    return results

# ── Main ───────────────────────────────────────────────────────────────────────
with open(CAPIQ_FILE, 'r', encoding='utf-8') as f:
    capiq = json.load(f)

companies = capiq.get('companies', capiq)
updated = 0

for fname in os.listdir(REPORTS_DIR):
    if not fname.endswith('.xlsx'): continue
    # Extract ticker from filename e.g. "EntergyCorporationNYSEETR_Report..."
    m = re.search(r'(?:NYSE|NASDAQGS|NASDAQGM|NASDAQCM|NASDAQ|NasdaqGS|NasdaqCM|OTC)([A-Z0-9]+)_Report', fname)
    if not m:
        continue
    ticker = m.group(1)

    if ticker not in companies: continue

    path = os.path.join(REPORTS_DIR, fname)
    try:
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    except Exception as e:
        print(f"  {ticker}: could not open ({e})")
        continue

    sheet = next((n for n in wb.sheetnames if n == 'Surprise'), None)
    if not sheet:
        sheet = next((n for n in wb.sheetnames if n.lower() == 'surprise'), None)
    if not sheet:
        print(f"  {ticker}: no Surprise sheet")
        continue

    ws = wb[sheet]
    surprise_data = extract_surprise_sheet(ws)
    if surprise_data:
        companies[ticker]['earnings_surprise'] = surprise_data
        quarters = [r for r in surprise_data if r['period_type']=='quarterly']
        annual   = [r for r in surprise_data if r['period_type']=='annual']
        print(f"  {ticker}: {len(quarters)} quarters, {len(annual)} annual periods")
        updated += 1
    else:
        print(f"  {ticker}: could not parse Surprise sheet")

# Write back
capiq['companies'] = companies
with open(CAPIQ_FILE, 'w', encoding='utf-8') as f:
    json.dump(capiq, f, indent=2)

print(f"\nDone. Updated {updated} companies.")