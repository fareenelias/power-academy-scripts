r"""
extract_all.py  —  Master CapIQ Excel Extractor
Power Academy | Weekly workflow

Drop new Excel reports into E:\PowerAcademy\data\reports\ then run:
    python E:\PowerAcademy\scripts\extract_all.py

Extracts per company (ONE workbook.open per file):
  'Pending Rate Cases' → capiq_export.json  pending_rate_cases
  'Past Rate Cases'    → capiq_export.json  past_rate_cases
  'Surprise'           → capiq_export.json  earnings_surprise
  'People'             → capiq_export.json  executives (name+title only)
                         executives_export.json  management/board (full bios)
  'Compensation'       → executives_export.json  salary / total_comp

Preserves all other fields in capiq_export.json unchanged
(financial time series, descriptions, NUP, customers, etc.)

Outputs:
  E:\PowerAcademy\data\capiq_export.json       updated in place
  E:\PowerAcademy\data\executives_export.json  full rebuild each run

Also (extended data layers, added this session):
  - Rips any new earnings-call transcripts via rip_earnings.py (incremental)
  - Health-checks earnings_calls.json (call-notes) and rra_states.json (RRA):
    coverage vs the 23-name universe, Q&A page-link coverage, deck flags,
    ripped-but-not-yet-built transcripts, and pending RRA placeholders.
  (Building call-notes entries from the ripped .txt stays a manual in-Claude step.)
"""

import os, re, glob, json
from datetime import datetime

try:
    import openpyxl
except ImportError:
    import subprocess, sys
    subprocess.check_call([sys.executable, '-m', 'pip', 'install', 'openpyxl',
                           '--break-system-packages', '-q'])
    import openpyxl

# ── Config ────────────────────────────────────────────────────────────────────

REPORTS_DIR   = r'E:\PowerAcademy\data\reports'
CAPIQ_JSON    = r'E:\PowerAcademy\data\capiq_export.json'
EXECS_JSON    = r'E:\PowerAcademy\data\executives_export.json'
TODAY         = datetime.now().strftime('%m-%d-%Y')

# ── Extended data layers (added: RRA + earnings call notes) ────────────────────
EARNINGS_JSON   = r'E:\PowerAcademy\data\earnings_calls.json'
RRA_JSON        = r'E:\PowerAcademy\data\rra_states.json'
TRANSCRIPTS_DIR = r'E:\PowerAcademy\Documents\transcripts'
RIP_SCRIPT      = r'E:\PowerAcademy\scripts\rip_earnings.py'

# 23-name coverage universe (for call-notes coverage reporting)
COVERAGE_UNIVERSE = [
    'NEE','D','ETR','CMS','PPL','AEE','POR','EIX','PCG','HE','EVRG','ES','VST','TLN','XIFR',
    'AWR','CWT','YORW','GWRS','AWK','WTRG','HTO','MSEX',
]

# Tickers that do NOT hold earnings calls (small water utilities) - excluded from
# call-notes coverage gaps rather than flagged as missing.
NO_CALL_TICKERS = {'YORW', 'MSEX'}


# ── Ticker extraction ─────────────────────────────────────────────────────────

def ticker_from_filename(fname):
    """
    Extract ticker from filenames like:
      'NextEraEnergy,Inc.NYSENEE_Report_06-28-2026.xlsx'
      'Evergy,Inc.NASDAQGSEVRG_Report_06-28-2026.xlsx'
    Longest prefixes must be listed first so NASDAQGS matches before NASDAQ.
    """
    base = re.sub(r'_Report_.*$', '', os.path.splitext(os.path.basename(fname))[0])
    for prefix in ['NASDAQGS', 'NASDAQGM', 'NASDAQCM', 'NASDAQ', 'NYSEARCA', 'NYSE', 'TSX', 'OTC']:
        if prefix in base:
            ticker = base.split(prefix)[-1].strip()
            if ticker:
                return ticker
    return None


# ── Shared helpers ────────────────────────────────────────────────────────────

def safe(v):
    """Return float, formatted date string, or None. Never raises."""
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.strftime('%m/%Y')
    s = str(v).strip()
    if s.lower() in ('na', 'n/a', 'none', '', '-', '--', '—'):
        return None
    try:
        return float(s.replace(',', ''))
    except ValueError:
        pass
    for fmt in ('%Y-%m-%d %H:%M:%S', '%Y-%m-%d', '%m/%d/%Y', '%m/%Y'):
        try:
            dt = datetime.strptime(s.split(' ')[0], fmt.split(' ')[0])
            return dt.strftime('%m/%Y')
        except ValueError:
            continue
    return s


def safe_str(v):
    if v is None:
        return None
    s = str(v).strip()
    return None if s.lower() in ('na', 'n/a', 'none', '', '-') else s


def safe_float(v):
    r = safe(v)
    if isinstance(r, float):
        return r
    if isinstance(r, str):
        try:
            return float(r.replace(',', '').replace('%', '').replace('$', ''))
        except ValueError:
            pass
    return None


def safe_date(v):
    r = safe(v)
    if r is None:
        return None
    if isinstance(r, str) and re.match(r'\d{2}/\d{4}', r):
        return r
    if isinstance(r, str):
        for fmt in ('%Y-%m-%d', '%m/%d/%Y', '%m/%Y', '%Y'):
            try:
                return datetime.strptime(r[:10] if len(r) > 10 else r, fmt).strftime('%m/%Y')
            except ValueError:
                continue
    return None


def clean(v):
    """Return stripped string or None for NA / empty / None."""
    if v is None:
        return None
    s = str(v).strip().replace('\xa0', ' ')
    return s if s and s.upper() not in ('NA', 'N/A', 'NONE', '') else None


def to_int(v):
    if v is None:
        return None
    try:
        n = int(str(v).strip())
        return n if n > 0 else None
    except Exception:
        return None


def comp_val(v):
    if v is None:
        return None
    s = str(v).strip()
    if s.upper() in ('NA', 'N/A', 'NONE', ''):
        return None
    try:
        return int(float(s.replace(',', '')))
    except Exception:
        return None


def norm_name(name):
    """Normalise exec name for compensation matching."""
    s = re.sub(r'"[^"]*"', '', str(name))
    s = re.sub(r'\s+(Jr\.?|Sr\.?|II|III|IV)\s*$', '', s, flags=re.I)
    s = re.sub(r'\xa0', ' ', s)
    return re.sub(r'\s+', ' ', s).strip().lower()


def ws_rows(ws, max_row=600):
    return list(ws.iter_rows(max_row=max_row, values_only=True))


# ── Rate Cases: Pending ───────────────────────────────────────────────────────
#
# Header at row 8 (0-indexed: 7):
#  0:State | 1:Company | 2:Ticker | 3:Docket | 4:Service | 5:CaseType |
#  6:FilingDate | 7:RevenueReq($M) | 8:RevChange% | 9:ROCCR% | 10:ROE% |
#  11:Equity% | 12:TestYear | 13:RateBase($M) | 14:ValuationMethod | 15:ExpectedDecision

PENDING_COLS = {
    'state': 0, 'company': 1, 'docket': 3, 'service_type': 4, 'case_type': 5,
    'filing_date': 6, 'req_revenue': 7, 'req_rev_pct': 8, 'req_roccr': 9,
    'req_roe': 10, 'req_equity': 11, 'test_year': 12, 'req_rate_base': 13,
    'rate_base_method': 14, 'decision_date': 15,
}

def extract_pending(ws):
    rows = ws_rows(ws)
    hdr_idx = None
    for i, row in enumerate(rows):
        vals = [str(v or '').lower() for v in row]
        if 'state' in vals and 'docket' in vals and 'expected decision date' in ' '.join(vals):
            hdr_idx = i
            break
    if hdr_idx is None:
        return []
    results = []
    for row in rows[hdr_idx + 1:]:
        if not any(v for v in row):
            continue
        state = safe_str(row[0] if len(row) > 0 else None)
        if not state:
            continue
        c = {}
        for field, col in PENDING_COLS.items():
            v = row[col] if col < len(row) else None
            if field in ('filing_date', 'test_year', 'decision_date'):
                c[field] = safe_date(v)
            elif field in ('req_revenue', 'req_rev_pct', 'req_roccr', 'req_roe',
                           'req_equity', 'req_rate_base'):
                c[field] = safe_float(v)
            else:
                c[field] = safe_str(v)
        if c.get('company') or c.get('docket'):
            results.append(c)
    return results


# ── Rate Cases: Past ──────────────────────────────────────────────────────────
#
# Two header rows at R5+R6 (indices 4+5). Data from R7 (index 6).
#  0:State | 1:Company | 3:Docket | 4:Service | 5:CaseType
#  6:FilingDate | 7:ReqRevenue | 8:ReqROCCR | 9:ReqROE | 10:ReqEquity | 11:ReqRateBase
#  12:DecisionDate | 13:DecisionType | 14:AuthRevenue | 15:PhaseIn | 16:Interim |
#  17:AuthROCCR | 18:AuthROE | 19:AuthEquity | 20:TestYear | 21:AuthRateBase
#  22:RateBaseMethod | 23:DurationMonths

PAST_COLS = {
    'state': 0, 'company': 1, 'docket': 3, 'service_type': 4, 'case_type': 5,
    'filing_date': 6, 'req_revenue': 7, 'req_roccr': 8, 'req_roe': 9,
    'req_equity': 10, 'req_rate_base': 11,
    'decision_date': 12, 'decision_type': 13, 'auth_revenue': 14,
    'phase_in': 15, 'interim': 16, 'auth_roccr': 17, 'auth_roe': 18,
    'auth_equity': 19, 'test_year': 20, 'auth_rate_base': 21,
    'rate_base_method': 22, 'duration_months': 23,
}

def extract_past(ws):
    rows = ws_rows(ws)
    data_start = None
    for i, row in enumerate(rows):
        vals = [str(v or '').lower() for v in row[:3]]
        if i >= 5 and vals[0] and vals[0] not in ('state', 'date', '') and len(vals[0]) > 1:
            data_start = i
            break
    if data_start is None:
        return []
    results = []
    for row in rows[data_start:]:
        if not any(v for v in row):
            continue
        state = safe_str(row[0] if len(row) > 0 else None)
        if not state or state.lower() in ('state', 'date'):
            continue
        c = {}
        for field, col in PAST_COLS.items():
            v = row[col] if col < len(row) else None
            if field in ('filing_date', 'decision_date', 'test_year'):
                c[field] = safe_date(v)
            elif field in ('req_revenue', 'req_roccr', 'req_roe', 'req_equity', 'req_rate_base',
                           'auth_revenue', 'auth_roccr', 'auth_roe', 'auth_equity', 'auth_rate_base',
                           'duration_months'):
                c[field] = safe_float(v)
            else:
                c[field] = safe_str(v)
        if c.get('company') or c.get('docket'):
            results.append(c)
    return results


# ── Earnings Surprise ─────────────────────────────────────────────────────────
#
# 'Surprise' sheet structure (confirmed for ETR, YORW):
#   Row N:   'Per Share Level' | 'FQ1 2025 - Mar 2025' | 'FQ2 ...' | 'FY ...'
#   Row N+x: 'EPS Normalized' (or 'EPS (GAAP)', 'EPS Diluted') — triggers block
#   Sub-rows (leading spaces): '    %', '    Difference($)', '    Actual ($)',
#                               '    Estimate($)', '    Announced Date'

def parse_period(label):
    s = str(label).strip()
    if s.startswith('FQ'):
        m = re.match(r'FQ(\d) (\d{4})', s)
        if m:
            return s, 'quarterly', int(m.group(1)), int(m.group(2))
    if s.startswith('FY'):
        m = re.match(r'FY (\d{4})', s)
        if m:
            return s, 'annual', None, int(m.group(1))
    return s, 'unknown', None, None


def extract_surprise(ws):
    rows = list(ws.iter_rows(max_row=80, values_only=True))

    # Find 'Per Share Level' header row
    header_idx = None
    for i, row in enumerate(rows):
        if row[0] and 'Per Share Level' in str(row[0]):
            header_idx = i
            break
    if header_idx is None:
        return []

    # Period columns from header row
    raw_periods = [(j + 1, v) for j, v in enumerate(rows[header_idx][1:])
                   if v and str(v).strip()]
    if not raw_periods:
        return []
    col_indices  = [c for c, _ in raw_periods]
    period_labels = [v for _, v in raw_periods]

    EPS_TRIGGERS = ('EPS Normalized', 'EPS (GAAP)', 'EPS Diluted')

    surprise_pct = {}
    surprise_d   = {}
    actual       = {}
    estimate     = {}
    announced    = {}
    in_eps       = False

    for row in rows[header_idx + 1:]:
        label   = str(row[0]).strip() if row[0] else ''
        stripped = label.lstrip()

        if any(k in stripped for k in EPS_TRIGGERS) and not label.startswith(' '):
            in_eps = True
            for ci, pl in zip(col_indices, period_labels):
                surprise_pct[pl] = safe_float(row[ci]) if ci < len(row) else None
            continue

        if not in_eps:
            continue

        if 'Difference' in stripped and '$' in stripped:
            for ci, pl in zip(col_indices, period_labels):
                surprise_d[pl] = safe_float(row[ci]) if ci < len(row) else None
        elif 'Actual' in stripped and '$' in stripped:
            for ci, pl in zip(col_indices, period_labels):
                actual[pl] = safe_float(row[ci]) if ci < len(row) else None
        elif 'Estimate' in stripped and '$' in stripped:
            for ci, pl in zip(col_indices, period_labels):
                estimate[pl] = safe_float(row[ci]) if ci < len(row) else None
        elif 'Announced' in stripped:
            for ci, pl in zip(col_indices, period_labels):
                announced[pl] = clean(row[ci]) if ci < len(row) else None
        elif stripped and not stripped.startswith(
                ('%', 'Difference', 'Actual', 'Estimate', 'Announced',
                 'Guidance', 'Accounting')):
            break  # new metric block — stop

    results = []
    for pl in period_labels:
        key, ptype, quarter, year = parse_period(pl)
        results.append({
            'period':           key,
            'period_type':      ptype,
            'quarter':          quarter,
            'year':             year,
            'eps_actual':       actual.get(pl),
            'eps_estimate':     estimate.get(pl),
            'eps_surprise_d':   surprise_d.get(pl),
            'eps_surprise_pct': surprise_pct.get(pl),
            'announced_date':   announced.get(pl),
        })
    return results


# ── People (Executives + Board) ───────────────────────────────────────────────
#
# 'People' sheet structure (confirmed):
#   Row: 'Top Executives'
#   Row: 'Name' | 'Role' | 'Status' | 'Age' | 'Gender' | 'Office' | 'Phone' | 'Email'
#        | '' | '' | 'LinkedIn' | 'Biography'        ← bio at col 11
#   ... exec rows ...
#   Row: 'Average'
#   Row: 'Board of Directors'
#   Row: 'Name' | 'Role' | 'Status' | 'Age' | 'Gender' | 'Begin Year' | 'End Year'
#        | 'Tenure' | 'Office' | 'Phone' | 'Email' | '' | '' | 'LinkedIn' | 'Biography'
#        ← bio at col 14, tenure at col 7, begin_year at col 5

STOP_SECTIONS = frozenset((
    'Other Board Members', 'Professionals',
    'Takeover Defenses', 'Latest Holders - Insiders/Strategic Owners',
))

def extract_people(ws):
    management, board = [], []
    mode                = None
    awaiting_exec_hdr  = False
    awaiting_board_hdr = False

    for row in ws.iter_rows(values_only=True):
        col0 = clean(row[0]) if row else None

        if col0 == 'Top Executives':
            awaiting_exec_hdr = True
            continue
        if col0 == 'Board of Directors':
            awaiting_exec_hdr  = False
            awaiting_board_hdr = True
            mode = None
            continue
        if col0 in STOP_SECTIONS:
            break
        if col0 == 'Average':
            mode = None
            continue
        if col0 == 'Name':
            if awaiting_exec_hdr:
                mode = 'exec'
                awaiting_exec_hdr = False
            elif awaiting_board_hdr:
                mode = 'board'
                awaiting_board_hdr = False
            continue
        if not col0:
            continue

        if mode == 'exec':
            management.append({
                'name':   clean(row[0]),
                'title':  clean(row[1]),
                'age':    to_int(row[3]),
                'gender': clean(row[4]),
                'email':  clean(row[7]),
                'bio':    clean(row[11]) if len(row) > 11 else None,
            })
        elif mode == 'board':
            board.append({
                'name':       clean(row[0]),
                'title':      clean(row[1]),
                'age':        to_int(row[3]),
                'gender':     clean(row[4]),
                'begin_year': to_int(row[5]),
                'end_year':   to_int(row[6]),
                'tenure':     clean(row[7]),
                'email':      clean(row[10]),
                'bio':        clean(row[14]) if len(row) > 14 else None,
            })

    return management, board


# ── Compensation ──────────────────────────────────────────────────────────────
#
# 'Compensation' sheet — 'Executive Compensation - Summary' section:
#   Period header row: blank col0, 'FY' in other cols
#   Exec block header: "Name - Title, Company"  (all other cols empty)
#   Data rows: 'Salary', 'As Reported Total Executive Compensation', etc.

def extract_compensation(ws):
    result      = {}
    in_summary  = False
    recent_col  = None
    current_key = None

    for row in ws.iter_rows(values_only=True):
        col0 = str(row[0]).strip() if row[0] is not None else ''

        if col0 == 'Executive Compensation - Summary':
            in_summary = True
            continue
        if not in_summary:
            continue

        # Period header row
        if not col0 and any(v and 'FY' in str(v) for v in row[1:]):
            period_cols = [i for i, v in enumerate(row) if v and 'FY' in str(v)]
            if period_cols:
                recent_col = period_cols[-1]
            continue

        # Exec block header: "Name - Title"
        if (col0 and ' - ' in col0
                and all((row[i] is None or str(row[i]).strip() == '')
                        for i in range(1, min(4, len(row))))):
            current_key = norm_name(col0.split(' - ')[0].strip())
            if current_key not in result:
                result[current_key] = {}
            continue

        if current_key is None or recent_col is None:
            continue

        v = row[recent_col] if recent_col < len(row) else None
        n = comp_val(v)
        if col0 == 'Salary':
            result[current_key]['salary'] = n
        elif col0 == 'As Reported Total Executive Compensation':
            result[current_key]['total_comp'] = n

    return result


def link_comp(management, comp):
    for person in management:
        key = norm_name(person['name'])
        c   = comp.get(key)
        if not c:
            last = key.split()[-1]
            for ck, cv in comp.items():
                if ck.split()[-1] == last and abs(len(ck) - len(key)) < 10:
                    c = cv
                    break
        person['salary_latest']     = c.get('salary')     if c else None
        person['total_comp_latest'] = c.get('total_comp') if c else None



# ── M&A History ───────────────────────────────────────────────────────────────
#
# 'Detailed M&A History' sheet structure (confirmed):
#   R1-R15: filter metadata (skip)
#   Rx:    '[Sector] Deals' — section group label (col0 only, col1+ empty)
#   Rx+1:  'Transaction ID' | 'Announcement Date' | 'Completion Date' | 'Target'
#           | 'Buyer' | 'Seller' | 'Role' | 'Transaction Value ($M)'
#           | 'Acquisition or Sale' | 'Transaction Type'
#   Rx+2+: data rows (col0 starts with 'SPTRD')
#   Repeated per sector section. Ends with footnote '* Indicates...'

def safe_date_or_status(v):
    """Return formatted date, 'Terminated', or None."""
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.strftime('%m/%Y')
    s = str(v).strip()
    if s.lower() in ('na', 'n/a', 'none', ''):
        return None
    if 'terminated' in s.lower() or 'withdrawn' in s.lower():
        return 'Terminated'
    return safe_date(v)


def safe_value_m(v):
    """Return float $M or None."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return round(float(v), 1) if v > 0 else None
    s = str(v).strip()
    if s.upper() in ('NA', 'N/A', '', 'NONE'):
        return None
    try:
        n = float(s.replace(',', ''))
        return round(n, 1) if n > 0 else None
    except Exception:
        return None


US_STATES = {
    'Alabama','Alaska','Arizona','Arkansas','California','Colorado','Connecticut',
    'Delaware','Florida','Georgia','Hawaii','Idaho','Illinois','Indiana','Iowa',
    'Kansas','Kentucky','Louisiana','Maine','Maryland','Massachusetts','Michigan',
    'Minnesota','Mississippi','Missouri','Montana','Nebraska','Nevada',
    'New Hampshire','New Jersey','New Mexico','New York','North Carolina',
    'North Dakota','Ohio','Oklahoma','Oregon','Pennsylvania','Rhode Island',
    'South Carolina','South Dakota','Tennessee','Texas','Utah','Vermont',
    'Virginia','Washington','West Virginia','Wisconsin','Wyoming',
}

GENERATION_INDUSTRIES = {
    'Renewable Electricity', 'Thermal Electricity', 'Nuclear Electricity',
    'Independent Power Producers & Energy Traders',
}

def parse_mw(name):
    if not name: return None
    m = re.search(r'(\d+(?:\.\d+)?)\s*[-]?\s*(?:MW|megawatt)', name, re.I)
    if m: return float(m.group(1))
    m = re.search(r'(\d+(?:\.\d+)?)\s*[-]?\s*GW', name, re.I)
    if m: return float(m.group(1)) * 1000
    return None

def parse_asset_count(name):
    if not name: return None
    words = {'two':2,'three':3,'four':4,'five':5,'six':6,'seven':7,'eight':8,'nine':9,'ten':10}
    first = name.split()[0].lower()
    if first in words: return words[first]
    m = re.match(r'^(\d+)\s+', name)
    if m:
        n = int(m.group(1))
        return n if n > 1 else None
    return None

def parse_asset_state(name):
    if not name: return None
    clean = name.rstrip('*').strip()
    for state in US_STATES:
        if clean.endswith(state): return state
    return None

METADATA_PREFIXES = (
    'Transaction Type:', 'Include Transaction', 'Private Equity',
    'Target/Issuer', 'Group By:', 'Geography:', 'Date Range:',
    'Date:', 'Transaction Value:', 'Transaction Status:',
    'Company Role:', 'Financials Source:', '* Indicates',
    'Acquisitions/Investments', 'Divestures', 'Shelf Registrations',
)

def extract_ma(ws):
    """
    Extract M&A deals from 'Detailed M&A History' sheet.
    Handles multiple sector sections, deduplicates by transaction_id.
    Returns list of deal dicts sorted newest-announced first.
    """
    deals      = []
    seen_ids   = set()
    sector     = None
    hdr_active = False

    for row in ws.iter_rows(values_only=True):
        col0 = str(row[0]).strip() if row[0] is not None else ''
        col1 = row[1] if len(row) > 1 else None

        # Skip metadata / filter header rows
        if any(col0.startswith(p) for p in METADATA_PREFIXES):
            continue

        # Skip fully empty rows
        if not col0 and not any(v for v in row):
            continue

        # Skip CapIQ company header rows ("Company Name | Sheet Name")
        if ' | ' in col0 and col1 is None:
            continue

        # Sector section label: ends with ' Deals', col1 is empty
        if col0.endswith(' Deals') and not col1:
            sector     = col0.replace(' Deals', '').strip()
            hdr_active = False
            continue

        # Column header row
        if col0 == 'Transaction ID':
            hdr_active = True
            continue

        # Data row: SPTRD prefix
        if not col0.startswith('SPTRD'):
            continue
        if not hdr_active:
            continue

        txn_id = col0
        if txn_id in seen_ids:
            continue
        seen_ids.add(txn_id)

        target   = safe_str(row[3] if len(row) > 3 else None)
        value_m  = safe_value_m(row[7] if len(row) > 7 else None)
        industry = safe_str(row[11] if len(row) > 11 else None)
        is_gen   = industry in GENERATION_INDUSTRIES

        capacity_mw   = parse_mw(target) if is_gen else None
        asset_state   = parse_asset_state(target)
        asset_count   = parse_asset_count(target)
        dollar_per_kw = round(value_m * 1000 / capacity_mw, 0) if (is_gen and capacity_mw and value_m) else None

        deals.append({
            'transaction_id':  txn_id,
            'announced':       safe_date(row[1] if len(row) > 1 else None),
            'completed':       safe_date_or_status(row[2] if len(row) > 2 else None),
            'target':          target,
            'buyer':           safe_str(row[4] if len(row) > 4 else None),
            'seller':          safe_str(row[5] if len(row) > 5 else None),
            'role':            safe_str(row[6] if len(row) > 6 else None),
            'value_m':         value_m,
            'acq_or_sale':     safe_str(row[8] if len(row) > 8 else None),
            'deal_type':       safe_str(row[9] if len(row) > 9 else None),
            'target_industry': industry,
            'is_generation':   is_gen,
            'capacity_mw':     capacity_mw,
            'asset_state':     asset_state,
            'asset_count':     asset_count,
            'dollar_per_kw':   dollar_per_kw,
            'sector':          sector,
            'advisors':        [],
        })

    return deals



def extract_advisors(ws):
    """
    Parse Advisers sheet M&A section.
    Returns {sptrd_id: ['Bank A', 'Bank B', ...]}
    Multi-bank deals have blank col0 continuation rows.
    """
    advisor_map = {}
    in_ma = False
    hdr_seen = False
    current_id = None

    for row in ws.iter_rows(values_only=True):
        col0 = str(row[0]).strip() if row[0] is not None else ''

        if col0 == 'M&A Advisers':
            in_ma = True; hdr_seen = False; continue
        if col0 == 'Offering Underwriters':
            break
        if not in_ma:
            continue
        if col0 == 'Transaction ID':
            hdr_seen = True; continue
        if not hdr_seen:
            continue

        bank = str(row[7]).strip() if len(row) > 7 and row[7] else ''
        bank = bank.strip()
        if bank.upper() in ('NA', 'N/A', ''):
            bank = None

        if col0.startswith('SPTRD'):
            current_id = col0
            if current_id not in advisor_map:
                advisor_map[current_id] = []
            if bank:
                advisor_map[current_id].append(bank)
        elif col0 == '' and current_id and bank:
            advisor_map[current_id].append(bank)

    return advisor_map


# ── Credit Sheets ─────────────────────────────────────────────────────────────
#
# Capital Structure Details: row 8 = header, data from row 9
#   0:Description | 1:Type | 2:Amount($K) | 3:Coupon | 4:FloatingRate
#   5:MaturityDate | 6:Seniority | 7:Secured
#
# Current Ratings: sections per agency, each with "Rating Type" header row
#   0:RatingType | 1:Rating | 2:RatingDate | 3:LastReview | 4:PrevRating
#   5:Action | 6:Outlook | 7:OutlookDate
#
# Credit Ratios (x): simple key-value, col0=metric, col1=value (current quarter)

def parse_maturity_year(v):
    """Extract 4-digit year from maturity date string."""
    if not v: return None
    s = str(v).strip()
    if s.upper() in ('NA','N/A',''): return None
    # Range like "2028 - 2065": use first year
    if ' - ' in s: s = s.split(' - ')[0].strip()
    m = re.search(r'\b(20\d{2})\b', s)
    return int(m.group(1)) if m else None


def extract_credit_details(ws):
    """Extract per-instrument debt from Capital Structure Details sheet."""
    instruments = []
    header_seen = False
    for row in ws.iter_rows(values_only=True):
        col0 = str(row[0]).strip() if row[0] else ''
        if col0 == 'Capital Structure Description':
            header_seen = True; continue
        if not header_seen or not col0: continue
        # Skip summary/footnote rows
        if col0.startswith('Total') or col0.startswith('*'): continue
        # Amount
        raw_amt = row[2] if len(row) > 2 else None
        try:
            amt_k = int(float(str(raw_amt).replace(',',''))) if raw_amt and str(raw_amt).strip().upper() not in ('NA','') else None
        except Exception:
            amt_k = None
        if not amt_k or amt_k <= 0: continue
        maturity_raw = str(row[5]).strip() if len(row) > 5 and row[5] else None
        if maturity_raw and str(maturity_raw).upper() in ('NA','N/A',''): maturity_raw = None
        # Infer issuing entity from description prefix or security type
        seniority = safe_str(row[6] if len(row) > 6 else None) or ''
        secured   = safe_str(row[7] if len(row) > 7 else None) or ''
        # First word may be a known abbreviation (SCE, DVP, AEE-MO, etc.)
        first_word = col0.split()[0] if col0 else ''
        if re.match(r'^[A-Z]{2,5}$', first_word) and len(col0.split()) > 1:
            inferred_entity = first_word
        elif 'First Mortgage' in col0 or ('Secured' in seniority and secured == 'Yes'):
            inferred_entity = 'OpCo (secured)'
        elif 'Junior Subordinated' in col0 or 'Junior Sub' in col0:
            inferred_entity = 'HoldCo (hybrid)'
        elif secured == 'No' and 'Senior' in seniority:
            inferred_entity = 'HoldCo (unsecured)'
        else:
            inferred_entity = None

        instruments.append({
            'description':    col0,
            'type':           safe_str(row[1] if len(row) > 1 else None),
            'amount_k':       amt_k,
            'coupon':         safe_str(row[3] if len(row) > 3 else None),
            'floating_rate':  safe_str(row[4] if len(row) > 4 else None),
            'maturity_raw':   maturity_raw,
            'maturity_year':  parse_maturity_year(maturity_raw),
            'seniority':      seniority or None,
            'secured':        secured or None,
            'issuing_entity': inferred_entity,
        })
    return instruments


def extract_current_ratings(ws):
    """Extract credit ratings by agency from Current Ratings sheet."""
    ratings = []
    current_agency = None
    current_entity = None
    header_seen    = False
    AGENCY_KEYS    = ('S&P GLOBAL', 'MOODY', 'FITCH', 'DBRS', 'KBRA')

    def fmt_dt(v):
        if not v: return None
        if isinstance(v, datetime): return v.strftime('%Y-%m-%d')
        s = str(v).strip()[:10]
        return s if re.match(r'\d{4}-\d{2}-\d{2}', s) else None

    for row in ws.iter_rows(values_only=True):
        col0 = str(row[0]).strip() if row[0] else ''
        if any(kw in col0.upper() for kw in AGENCY_KEYS):
            current_agency = col0.split('(')[0].strip()
            # Entity name may be split: col0="...Entity Name:" col1="Dominion Energy Inc.)"
            col1_str = str(row[1]).strip() if len(row) > 1 and row[1] else ''
            full_line = col0 + (' ' + col1_str if col1_str else '')
            em = re.search(r'Entity Name:\s*([^)]+)', full_line)
            current_entity = em.group(1).strip().rstrip('.)') if em else None
            header_seen = False; continue
        if col0 == 'Rating Type':
            header_seen = True; continue
        if not header_seen or not col0: continue
        if col0 in ('Issuer Credit Rating',): continue
        rating_val = str(row[1]).strip() if len(row) > 1 and row[1] else None
        if not rating_val or rating_val in ('',' '): continue
        ratings.append({
            'agency':      current_agency or 'Unknown',
            'entity':      current_entity,
            'rating_type': col0,
            'rating':      rating_val,
            'rating_date': fmt_dt(row[2] if len(row) > 2 else None),
            'last_review': fmt_dt(row[3] if len(row) > 3 else None),
            'prev_rating': safe_str(row[4] if len(row) > 4 else None),
            'action':      safe_str(row[5] if len(row) > 5 else None),
            'outlook':     safe_str(row[6] if len(row) > 6 else None),
        })
    return ratings



def extract_debt_analysis(ws):
    """
    Extract interest coverage and coverage ratios from Debt Analysis sheet.
    Returns dict with most-recent-year values.
    Coverage Ratios section has header 'Coverage Ratios (x)' then key-value time-series rows.
    Period cols: 2021 2022 2023 2024 2025 — rightmost non-None = most recent.
    """
    result = {}
    in_coverage = False
    recent_col  = None

    for row in ws.iter_rows(values_only=True):
        col0 = str(row[0]).strip() if row[0] else ''

        # Find period header row: col0 blank, other cols have year-like values
        # Handles 'Dec 2025', '12/2025 FY', 'FY 2025', '2025' etc.
        if not col0 and any(v and re.search(r'20\d{2}', str(v)) for v in row[1:]):
            period_cols = [i for i,v in enumerate(row)
                           if v and re.search(r'20\d{2}', str(v))]
            if period_cols: recent_col = period_cols[-1]
            continue

        if 'coverage ratio' in col0.lower() or col0.lower() == 'coverage ratios (x)':
            in_coverage = True; continue

        if col0 == 'Debt Detail - Current ($000)':
            break  # past coverage section

        if not in_coverage or not col0 or recent_col is None:
            continue

        v = row[recent_col] if recent_col < len(row) else None
        try:
            val = float(str(v)) if v and str(v).strip().upper() not in ('NA','NM','') else None
        except Exception:
            val = None

        if col0 == 'Pre-tax Interest Coverage Excl. AFUDC':
            result['interest_coverage']      = round(val, 2) if val else None
        elif col0 == 'Recurring EBITDA/ Adjusted Interest Expense':
            result['ebitda_interest_coverage']= round(val, 2) if val else None
        elif col0 == 'Debt/ Recurring EBITDA':
            result['debt_ebitda_snl']         = round(val, 2) if val else None

    return result

def extract_credit_ratios(ws):
    """Extract pre-computed ratios from Credit Ratios (x) sheet."""
    ratios = {}
    for row in ws.iter_rows(max_row=20, values_only=True):
        col0 = str(row[0]).strip() if row[0] else ''
        col1 = row[1] if len(row) > 1 else None
        if not col0 or col1 is None: continue
        try:
            val = float(str(col1).strip())
            ratios[col0] = round(val, 3)
        except Exception:
            pass
    return ratios

# ── Main ──────────────────────────────────────────────────────────────────────

# ── Extended layer: earnings call notes ───────────────────────────────────────

def rip_transcripts():
    """Incremental transcript rip: convert any newly-dropped earnings-call PDFs
    to text (the source the call-notes entries are built from). No-op if the
    ripper or transcripts folder isn't present. Building the structured entries
    in earnings_calls.json from the resulting .txt remains a manual (in-Claude) step."""
    if not os.path.isfile(RIP_SCRIPT) or not os.path.isdir(TRANSCRIPTS_DIR):
        print('  [call notes] ripper or transcripts dir not found — skipping rip step')
        return
    import subprocess, sys
    print(f'  [call notes] ripping new/changed transcripts under {TRANSCRIPTS_DIR} ...')
    try:
        subprocess.run([sys.executable, RIP_SCRIPT, TRANSCRIPTS_DIR], check=False)
    except Exception as e:
        print(f'  [call notes] rip step failed: {e}')


def report_call_notes():
    """Health-check earnings_calls.json: coverage vs the 23-name universe, call
    counts, Q&A page-link coverage, deck-sourced flags, and any transcripts that
    have been ripped but not yet built into entries."""
    if not os.path.isfile(EARNINGS_JSON):
        print('  [call notes] earnings_calls.json not found — skipping')
        return
    try:
        data = json.load(open(EARNINGS_JSON, encoding='utf-8'))
    except Exception as e:
        print(f'  [call notes] could not read earnings_calls.json: {e}')
        return
    calls = data.get('calls', {})
    total = sum(len(v) for v in calls.values())
    qa_total = qa_nopage = 0
    deck_flags = []
    for tk, cc in calls.items():
        for c in cc:
            if c.get('_qc'):
                deck_flags.append(f"{tk} {c.get('period','?')}")
            for q in c.get('qa_highlights', []):
                qa_total += 1
                if not q.get('source_page'):
                    qa_nopage += 1
    callable_universe = [t for t in COVERAGE_UNIVERSE if t not in NO_CALL_TICKERS]
    covered = [t for t in callable_universe if t in calls]
    missing = [t for t in callable_universe if t not in calls]
    print(f'  [call notes] {total} calls across {len(calls)} tickers  '
          f'({len(covered)}/{len(callable_universe)} of call-holding universe)')
    print(f'               Q&A items {qa_total}  (missing page link: {qa_nopage})')
    if missing:
        print(f'               no notes yet: {", ".join(missing)}')
    if NO_CALL_TICKERS:
        print(f'               excluded (no earnings calls): {", ".join(sorted(NO_CALL_TICKERS))}')
    if deck_flags:
        print(f'               deck-sourced (re-rip transcript): {", ".join(sorted(set(deck_flags)))}')

    # ripped-but-not-built: compare the ripper index against what is in earnings_calls.json
    idx_path = os.path.join(TRANSCRIPTS_DIR, 'text', '_earnings_index.json')
    if os.path.isfile(idx_path):
        try:
            idx = json.load(open(idx_path, encoding='utf-8'))
            built = {(tk, c.get('period')) for tk, cc in calls.items() for c in cc}
            unbuilt = set()
            for r in idx:
                tk, per = r.get('ticker'), r.get('period')
                if not tk or tk == 'UNK' or per in (None, 'ND'):
                    continue
                if r.get('call_type') == 'special':
                    continue
                if r.get('flag') in ('NO_QA_STRUCTURE_LIKELY_DECK', 'IMAGE_ONLY_NO_OCR'):
                    continue
                if (tk, per) not in built:
                    unbuilt.add(f'{tk} {per}')
            if unbuilt:
                print(f'               ripped, not yet built ({len(unbuilt)}): '
                      + ', '.join(sorted(unbuilt)))
        except Exception:
            pass


# ── Extended layer: RRA state regulatory database ─────────────────────────────

def report_rra():
    """Health-check rra_states.json: number of state profiles and any categories
    still marked pending/placeholder (e.g. the large_load_tariff placeholders)."""
    if not os.path.isfile(RRA_JSON):
        print('  [RRA] rra_states.json not found — skipping')
        return
    try:
        rra = json.load(open(RRA_JSON, encoding='utf-8'))
    except Exception as e:
        print(f'  [RRA] could not read rra_states.json: {e}')
        return
    states = {k: v for k, v in rra.items() if not k.startswith('_')}
    placeholder_states = []
    for st, prof in states.items():
        try:
            blob = json.dumps(prof, ensure_ascii=False).lower()
        except Exception:
            continue
        if '\u23f3' in blob or '⏳' in blob or 'placeholder' in blob or '"pending"' in blob:
            placeholder_states.append(st)
    print(f'  [RRA] {len(states)} state profiles'
          + (f'  ({len(placeholder_states)} with pending/placeholder categories)'
             if placeholder_states else '  (all categories populated)'))


def main():
    # ── Load existing capiq_export.json ──────────────────────────────────────
    if os.path.exists(CAPIQ_JSON):
        with open(CAPIQ_JSON, 'r', encoding='utf-8') as f:
            capiq = json.load(f)
        companies = capiq.get('companies', capiq)
        print(f'Loaded {CAPIQ_JSON}: {len(companies)} existing companies')
    else:
        companies = {}
        capiq = {'companies': companies}
        print('Starting fresh capiq_export.json')

    # ── Find Excel files ──────────────────────────────────────────────────────
    files = sorted(glob.glob(os.path.join(REPORTS_DIR, '*.xlsx')))
    files = [f for f in files if not os.path.basename(f).startswith('~$')]
    print(f'\nProcessing {len(files)} Excel files from {REPORTS_DIR}\n')

    # Output containers
    execs_output = {}

    # Stats
    stats = {
        'pending': 0, 'past': 0,
        'surprise': 0, 'surprise_empty': 0,
        'mgmt': 0, 'board': 0,
        'ma': 0,
        'skipped': 0,
    }

    # ── Per-file extraction ───────────────────────────────────────────────────
    for fpath in files:
        ticker = ticker_from_filename(fpath)
        if not ticker:
            fname = os.path.basename(fpath)
            print(f'  [SKIP] No ticker found: {fname}')
            stats['skipped'] += 1
            continue

        fname_short = os.path.basename(fpath)[:52]
        print(f'  {ticker:<6} {fname_short}')

        try:
            wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
        except Exception as e:
            print(f'         ✗ Could not open: {e}')
            stats['skipped'] += 1
            continue

        sheets = set(wb.sheetnames)

        # Ensure ticker exists in companies dict
        if ticker not in companies:
            companies[ticker] = {'ticker': ticker, 'extracted': TODAY}

        # ── Rate Cases ───────────────────────────────────────────────────────
        pending, past = [], []
        if 'Pending Rate Cases' in sheets:
            pending = extract_pending(wb['Pending Rate Cases'])
        if 'Past Rate Cases' in sheets:
            past    = extract_past(wb['Past Rate Cases'])

        companies[ticker]['pending_rate_cases'] = pending
        companies[ticker]['past_rate_cases']    = past
        stats['pending'] += len(pending)
        stats['past']    += len(past)
        print(f'         rate cases  pending={len(pending):2d}  past={len(past):3d}')

        # ── Earnings Surprise ─────────────────────────────────────────────────
        surprise_data = []
        if 'Surprise' in sheets:
            surprise_data = extract_surprise(wb['Surprise'])

        if surprise_data:
            companies[ticker]['earnings_surprise'] = surprise_data
            q = sum(1 for r in surprise_data if r['period_type'] == 'quarterly')
            a = sum(1 for r in surprise_data if r['period_type'] == 'annual')
            has_data = any(r['eps_actual'] is not None for r in surprise_data)
            print(f'         surprise    {q}Q {a}A  {"has actuals" if has_data else "all null — check sheet"}')
            stats['surprise'] += 1
        else:
            print(f'         surprise    no Surprise sheet or empty')
            stats['surprise_empty'] += 1

        # ── People / Executives ───────────────────────────────────────────────
        management, board = [], []
        if 'People' in sheets:
            management, board = extract_people(wb['People'])

        # Compensation
        comp = {}
        if 'Compensation' in sheets:
            comp = extract_compensation(wb['Compensation'])

        link_comp(management, comp)

        # Write simplified exec list to capiq_export.json (name + title only)
        companies[ticker]['executives'] = [
            {'name': p['name'], 'title': p['title']}
            for p in management + board
            if p.get('name')
        ]

        # Full detail to executives_export.json
        execs_output[ticker] = {
            'management': management,
            'board':      board,
            'source':     'S&P Capital IQ',
            'as_of':      TODAY,
        }

        comp_count = sum(1 for p in management if p.get('total_comp_latest'))
        print(f'         people      {len(management):2d} mgmt  {len(board):2d} board  '
              f'{comp_count}/{max(len(management),1)} have comp data')
        stats['mgmt']  += len(management)
        stats['board'] += len(board)

        # ── M&A History ──────────────────────────────────────────────────────
        ma_deals = []
        if 'Detailed M&A History' in sheets:
            ma_deals = extract_ma(wb['Detailed M&A History'])
        if 'Advisers' in sheets:
            adv_map = extract_advisors(wb['Advisers'])
            for deal in ma_deals:
                deal['advisors'] = adv_map.get(deal['transaction_id'], [])
        companies[ticker]['ma_history'] = ma_deals
        stats['ma'] += len(ma_deals)
        gen_ct = sum(1 for d in ma_deals if d.get('is_generation'))
        adv_ct = sum(1 for d in ma_deals if d.get('advisors'))
        print(f'         m&a         {len(ma_deals):2d} deals ({gen_ct} gen  {adv_ct} w/banks)')

        # ── Credit sheets ────────────────────────────────────────────────────
        cap_details, cur_ratings, cred_ratios, debt_analysis = [], [], {}, {}
        if 'Capital Structure Details' in sheets:
            cap_details  = extract_credit_details(wb['Capital Structure Details'])
        if 'Current Ratings' in sheets:
            cur_ratings  = extract_current_ratings(wb['Current Ratings'])
        if 'Credit Ratios (x)' in sheets:
            cred_ratios   = extract_credit_ratios(wb['Credit Ratios (x)'])
        if 'Debt Analysis' in sheets:
            debt_analysis = extract_debt_analysis(wb['Debt Analysis'])
        companies[ticker]['capital_structure_details'] = cap_details
        companies[ticker]['current_ratings']           = cur_ratings
        companies[ticker]['credit_ratios']             = cred_ratios
        companies[ticker]['debt_analysis']             = debt_analysis
        print(f'         credit      {len(cap_details):2d} instruments  {len(cur_ratings):2d} ratings  {len(cred_ratios):2d} ratios')

        wb.close()

    # ── Write outputs ─────────────────────────────────────────────────────────
    capiq['companies'] = companies
    with open(CAPIQ_JSON, 'w', encoding='utf-8') as f:
        json.dump(capiq, f, indent=2, ensure_ascii=False)

    with open(EXECS_JSON, 'w', encoding='utf-8') as f:
        json.dump(execs_output, f, indent=2, ensure_ascii=False)

    # ── Summary ───────────────────────────────────────────────────────────────
    print(f'\n{"="*62}')
    print(f'  Wrote: {CAPIQ_JSON}')
    print(f'         {EXECS_JSON}')
    print(f'{"="*62}')
    print(f'  Rate cases   pending={stats["pending"]}  past={stats["past"]}')
    print(f'  Surprise     {stats["surprise"]} updated  {stats["surprise_empty"]} no data')
    print(f'  People       {stats["mgmt"]} mgmt  {stats["board"]} board across {len(execs_output)} companies')
    print(f'  M&A          {stats["ma"]} deals across {sum(1 for c in companies.values() if c.get("ma_history"))} companies')
    if stats['skipped']:
        print(f'  Skipped      {stats["skipped"]} files')
    print()

    # Ticker summary table
    print(f'  {"Ticker":<6}  {"Pending":>7}  {"Past":>5}  {"Surprise":>8}  '
          f'{"Mgmt":>4}  {"Board":>5}  {"M&A":>5}')
    print(f'  {"-"*6}  {"-"*7}  {"-"*5}  {"-"*8}  {"-"*4}  {"-"*5}  {"-"*5}')
    for ticker in sorted(execs_output.keys()):
        co  = companies.get(ticker, {})
        pen = len(co.get('pending_rate_cases', []))
        pst = len(co.get('past_rate_cases', []))
        srp = len(co.get('earnings_surprise', []))
        srp_str = f'{srp}p' if srp else '—'
        mg  = len(execs_output[ticker]['management'])
        bd  = len(execs_output[ticker]['board'])
        ma  = len(co.get('ma_history', []))
        ma_str = str(ma) if ma else '—'
        print(f'  {ticker:<6}  {pen:>7}  {pst:>5}  {srp_str:>8}  {mg:>4}  {bd:>5}  {ma_str:>5}')

    # ── Extended data layers (RRA + call notes) ────────────────────────────────
    print(f'\n{"─"*62}')
    print('  Extended data layers')
    print(f'{"─"*62}')
    rip_transcripts()
    report_call_notes()
    report_rra()

    print(f'\n[✓] Done — {TODAY}')
    print('\nNext step:')
    print('  git add -A; git commit -m "data: weekly CapIQ extract ' + TODAY + '"')


if __name__ == '__main__':
    main()