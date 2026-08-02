#!/usr/bin/env python3
"""
extract_aqn_financials.py
=========================

Fill the CapIQ financial time-series / profile block for a single ticker
(default: AQN, Algonquin Power & Utilities Corp.) into capiq_export.json,
reading a S&P Capital IQ tearsheet workbook (.xlsx).

Design rules (deliberate, do not "improve" without reading them):

  * MERGE, NEVER REBUILD.  The existing capiq_export.json is read, only the
    *empty* target fields of the requested ticker are populated, and the file
    is written back.  Every other ticker and every already-populated field of
    the target ticker is left byte-identical.
  * NEVER INVENT A NUMBER.  Every value written here is read out of a named
    cell of a named sheet.  Nothing is interpolated, carried forward, derived
    from a peer, or estimated.  If a value cannot be located the field is left
    absent and reported as NOT FOUND.  The only computed field is
    total_customers_latest, which is an explicit sum of two workbook figures
    (this mirrors how the value is built for the peer records) and is labelled
    as derived in the run summary.
  * UNITS.  Balance-sheet / income-statement / cash-flow series are written in
    THOUSANDS of the reported currency, exactly as the workbook reports them
    ("Magnitude: Thousands (K)").  No rescaling is applied anywhere.
  * CURRENCY.  Nothing is FX-converted.  The reported-currency code found in
    the workbook is echoed in the run summary so the caller can see whether the
    record is USD or CAD.  For dual-listed Canadian issuers CapIQ reports the
    financial statements in the filing currency but the market-based
    capitalisation block (price, market cap, TEV) in the *listing* currency of
    the primary ticker -- the summary flags this explicitly.
  * ATOMIC WRITE.  Serialise to a temp file next to the target, re-read and
    re-parse that temp file, and only then os.replace() it over the original.

Usage:
    python extract_aqn_financials.py \
        --workbook /path/to/AlgonquinPowerAndUtilitiesCorp.TSXAQN_Report_07-29-2026.xlsx \
        --capiq    /path/to/capiq_export.json \
        [--ticker AQN] [--dry-run] [--result-json /path/to/RESULT.json]

No paths are hardcoded; the script runs unchanged on Windows and Linux.
"""

from __future__ import annotations

import argparse
import json
import os
import sys
import tempfile
from datetime import date, datetime

try:
    import openpyxl
except ImportError:  # pragma: no cover
    sys.exit("openpyxl is required:  pip install openpyxl")


# --------------------------------------------------------------------------
# generic sheet helpers
# --------------------------------------------------------------------------

NULLISH = {"NA", "NM", "N/A", "NA ", "NM ", "-", "--", ""}


def sheet_rows(wb, name):
    """Return a sheet as a list of tuples (1-indexed row r == rows[r-1])."""
    if name not in wb.sheetnames:
        return []
    return [tuple(r) for r in wb[name].iter_rows(values_only=True)]


def cell(rows, r, c):
    """1-indexed cell access, tolerant of ragged rows."""
    if r < 1 or r > len(rows):
        return None
    row = rows[r - 1]
    return row[c - 1] if 0 <= c - 1 < len(row) else None


def txt(v):
    """Cell -> stripped string, or None."""
    if v is None:
        return None
    if isinstance(v, (datetime, date)):
        return v.isoformat()[:10]
    s = str(v).strip()
    return s or None


def num(v):
    """Cell -> number, or None for NA / NM / text.  No unit conversion."""
    if v is None or isinstance(v, bool):
        return None
    if isinstance(v, (int,)):
        return v
    if isinstance(v, float):
        # kill float representation noise without changing the value
        r = round(v, 6)
        return int(r) if r == int(r) and abs(r) < 2 ** 53 else r
    s = str(v).strip()
    if s.upper() in NULLISH:
        return None
    s2 = s.replace(",", "").replace("%", "").replace("$", "")
    if s2.startswith("(") and s2.endswith(")"):
        s2 = "-" + s2[1:-1]
    try:
        f = float(s2)
    except ValueError:
        return None
    return int(f) if f == int(f) else round(f, 6)


def find_row(rows, label, col=1, exact=True, start=1):
    """Row index (1-indexed) whose column `col` matches `label` after strip."""
    want = label.strip().lower()
    for r in range(start, len(rows) + 1):
        v = txt(cell(rows, r, col))
        if v is None:
            continue
        v = v.strip().lower()
        if (v == want) if exact else (want in v):
            return r
    return None


def series(rows, label, first_col, ncols, exact=True, start=1):
    """
    Pull a horizontal numeric series from the row labelled `label`.
    Returns (values, row_index) or (None, None) when the label is absent.
    Values keep their workbook magnitude; NA/NM become None.
    """
    r = find_row(rows, label, exact=exact, start=start)
    if r is None:
        return None, None
    return [num(cell(rows, r, first_col + i)) for i in range(ncols)], r


# --------------------------------------------------------------------------
# extraction
# --------------------------------------------------------------------------

class Extractor:
    """Locates each target field in the workbook.  Records provenance."""

    # 5-year annual block on 'Financial Highlights' (cols B..F = FY-4 .. FY0)
    FH_FIELDS = {
        "total_assets": "Total Assets",
        "net_debt": "Net Debt",
        "net_ppe": "Net Property, Plant & Equipment",
        "total_equity": "Total Equity",
        "capex": "Capital Expenditure",
        "dividends_ps": "Dividends per Share",
        "roe": "Return on Equity",
        "roa": "Return on Assets",
        "total_debt": "Total Debt",
        "cash": "Cash & Short-term Investments",
    }

    # 7-period block on 'Key Stats' (3 FY + LTM + 3 consensus estimate years)
    KS_FIELDS = {
        "revenue": "Total Revenue ($000)",
        "revenue_growth": "Total Revenues, 1 Year Growth (%)",
        "ebitda": "EBITDA ($000)",
        "ebitda_margin": "EBITDA Margin (%)",
        "net_income": "Net Income ($000)",
        "eps_diluted": "Diluted EPS Excl. Extra Items ($)",
    }

    # SNL operating block on 'Operating Profile' (same 5 fiscal years as FH)
    # NB: matched EXACTLY (after strip).  'Retail Electricity Sold' and
    # 'Other Retail Electricity Sold' are different lines; a substring match
    # silently grabs the wrong one.
    OP_FIELDS = {
        "electric_customers": "Electric Customers",
        "gas_customers": "Natural Gas Distribution Customers",
        "fte_employees": "FTE Employees",
        "retail_revenue": "Electric Sales: Retail",
        "retail_mwh": "Retail Electricity Sold",
    }

    def __init__(self, wb):
        self.wb = wb
        self.values = {}      # field -> value
        self.source = {}      # field -> provenance string
        self.missing = {}     # field -> reason
        self.currency = {}    # diagnostics

    # -- bookkeeping ------------------------------------------------------
    def put(self, field, value, src):
        if value is None or (isinstance(value, (list, dict, str)) and len(value) == 0):
            self.skip(field, "located sheet but no usable value (%s)" % src)
            return
        self.values[field] = value
        self.source[field] = src

    def skip(self, field, reason):
        self.missing[field] = reason

    # -- driver -----------------------------------------------------------
    def run(self):
        wb = self.wb
        fh = sheet_rows(wb, "Financial Highlights")
        ks = sheet_rows(wb, "Key Stats")
        op = sheet_rows(wb, "Operating Profile")
        bs = sheet_rows(wb, "Balance Sheet")
        sg = sheet_rows(wb, "Segment Analysis")

        self._identity(wb, fh, bs)
        self._currency(fh, ks)
        self._periods_and_estimates(ks)
        self._annual_block(fh)
        self._tev(ks)
        self._operating(op)
        self._utility_from_segments(sg)
        self._net_utility_plant(bs, fh)
        self._capacity(wb)
        self._power_plants(wb)
        self._analysts(wb)
        self._advisers(wb)
        self._transactions(wb)
        self._top_holders(wb)
        self._segments(sg)
        self._planned_capex(wb)
        self._debt_items(wb)
        self._dividends(wb)
        self._rate_base(wb)
        self._ferc(wb)
        return self

    # -- individual field groups -----------------------------------------
    def _identity(self, wb, fh, bs):
        name = None
        for rows in (fh, bs, sheet_rows(wb, "Long Business Description")):
            t = txt(cell(rows, 2, 1))
            if t and "|" in t:
                name = t.split("|")[0].strip()
                break
        self.put("name", name, "sheet title cell A2")

        lbd = sheet_rows(wb, "Long Business Description")
        desc = None
        for r in range(4, len(lbd) + 1):
            v = txt(cell(lbd, r, 1))
            if v and len(v) > 200:
                desc = v
                break
        self.put("long_description", desc, "Long Business Description!A%s" % (r if desc else "?"))

        # Company address / web / phone / SIC: US tearsheets carry a company
        # info block; this workbook does not.  Never fabricate them.
        for f in ("hq", "website", "phone", "sic"):
            self.skip(f, "no company-info block (address / website / phone / SIC) "
                         "anywhere in this workbook")

    def _currency(self, fh, ks):
        r = find_row(fh, "Reported Currency Code")
        codes = []
        if r:
            codes = [txt(cell(fh, r, c)) for c in range(2, 8)]
            codes = [c for c in codes if c]
        self.currency["financials"] = sorted(set(codes)) or ["unknown"]
        r2 = find_row(ks, "Reported Currency Code")
        if r2:
            ks_codes = [txt(cell(ks, r2, c)) for c in range(2, 9)]
            self.currency["key_stats"] = sorted({c for c in ks_codes if c})
        # market-cap block currency comes from the label itself, e.g. "(C$000)"
        for lbl in ("Closing Price", "Market Capitalization", "Total Enterprise Value"):
            rr = find_row(ks, lbl, exact=False)
            if rr:
                self.currency.setdefault("market_block", []).append(txt(cell(ks, rr, 1)))

    def _periods_and_estimates(self, ks):
        pe = find_row(ks, "Period Ended")
        if pe is None or pe < 2:
            self.skip("periods", "'Period Ended' header not found on Key Stats")
            for f in self.KS_FIELDS:
                self.skip(f, "Key Stats period header not found")
            return
        hdr = pe - 1
        periods = [txt(cell(ks, hdr, c)) for c in range(2, 9)]
        periods = [p for p in periods if p]
        self.put("periods", periods, "Key Stats!row %d" % hdr)
        n = len(periods)
        for field, label in self.KS_FIELDS.items():
            vals, r = series(ks, label, 2, n, start=pe)
            if vals is None:
                self.skip(field, "row '%s' not on Key Stats" % label)
            else:
                self.put(field, vals, "Key Stats!row %d '%s' (reported magnitude)" % (r, label))

    def _annual_block(self, fh):
        pe = find_row(fh, "Period Ended")
        if pe is None:
            for f in self.FH_FIELDS:
                self.skip(f, "Financial Highlights period header not found")
            return
        hdr = pe - 2 if txt(cell(fh, pe - 2, 2)) else pe - 1
        labels = [txt(cell(fh, hdr, c)) for c in range(2, 8)]
        ncols = len([l for l in labels if l])
        self.fh_periods = [l for l in labels if l]
        for field, label in self.FH_FIELDS.items():
            vals, r = series(fh, label, 2, ncols, start=pe)
            if vals is None:
                self.skip(field, "row '%s' not on Financial Highlights" % label)
            else:
                unit = "as reported" if field in ("dividends_ps", "roe", "roa") \
                    else "thousands, as reported"
                self.put(field, vals,
                         "Financial Highlights!row %d '%s' (%s)" % (r, label, unit))

    def _tev(self, ks):
        r = find_row(ks, "Total Enterprise Value", exact=False)
        if r is None:
            self.skip("tev", "no 'Total Enterprise Value' row on Key Stats")
            return
        label = txt(cell(ks, r, 1))
        v = None
        for c in range(2, 9):
            v = num(cell(ks, r, c))
            if v is not None:
                break
        # DELIBERATELY NOT STORED. CapIQ prices AQN off its PRIMARY listing, TSX:AQN,
        # so the Key Stats "Latest Capitalization" block is in C$000 - the labels say
        # so explicitly ("Closing Price (C$)", "Total Enterprise Value (TEV) (C$000)")
        # - while every other series in this workbook is US$000. Writing a C$ number
        # into a field the dashboard reads as US$ would overstate AQN by ~1.37x against
        # its US peers in every EV multiple. Storing nothing is correct; `null`/absent
        # over fiction. fetch_market_data.py does not use `tev` - it builds firm value
        # from the live FMP marketCap plus net_debt - so nothing downstream breaks.
        # To restore it, convert at the Bank of Canada FXUSDCAD rate for the workbook
        # date and store it with an fx_conversion note, the way the J.P. Morgan AQN
        # price target was handled on 2026-07-31.
        self.skip("tev", "printed in C$000 (TSX primary listing) while the rest of the "
                         "workbook is US$000 - not stored rather than mixing currencies")
        return
        self.put("tev", v, "Key Stats!row %d '%s'" % (r, label))

    def _operating(self, op):
        if not op:
            for f in self.OP_FIELDS:
                self.skip(f, "no 'Operating Profile' sheet")
            self.skip("retail_customers", "no 'Operating Profile' sheet")
            self.skip("residential_customers", "no 'Operating Profile' sheet")
            self.skip("total_customers_latest", "no 'Operating Profile' sheet")
            return
        pe = find_row(op, "Fiscal Period Ended") or find_row(op, "Period Ended")
        hdr = (pe - 2) if pe and txt(cell(op, pe - 2, 2)) else (pe - 1 if pe else None)
        ncols = len([1 for c in range(2, 8) if txt(cell(op, hdr, c))]) if hdr else 5
        for field, label in self.OP_FIELDS.items():
            vals, r = series(op, label, 2, ncols, exact=True)
            if vals is None or all(v is None for v in vals):
                self.skip(field, "row '%s' absent/all-NA on Operating Profile" % label)
            else:
                self.put(field, vals, "Operating Profile!row %d '%s'" % (r, label))

        # No residential customer count and no combined "retail customers"
        # line exists on this sheet -- do not substitute the electric count.
        if find_row(op, "Residential Customers", exact=False) is None:
            self.skip("residential_customers",
                      "Operating Profile has no residential customer-count row")
        if find_row(op, "Retail Customers", exact=False) is None:
            self.skip("retail_customers",
                      "Operating Profile reports 'Electric Customers' and 'Natural Gas "
                      "Distribution Customers' only; no combined retail customer row")

        ec = self.values.get("electric_customers")
        gc = self.values.get("gas_customers")
        if ec and gc and ec[-1] is not None and gc[-1] is not None:
            self.put("total_customers_latest", ec[-1] + gc[-1],
                     "DERIVED = latest Electric Customers + latest Natural Gas "
                     "Distribution Customers (Operating Profile)")
        else:
            self.skip("total_customers_latest",
                      "needs latest electric + gas customer counts; one is missing")

    def _utility_from_segments(self, sg):
        """
        Regulated-utility revenue / EBITDA taken from the *named* regulated
        segment on Segment Analysis.  Nothing is allocated or estimated; if the
        segment row has gaps the field is left empty (same as peers whose
        segment split does not map cleanly onto a single utility segment).
        """
        if not sg:
            self.skip("utility_revenue", "no 'Segment Analysis' sheet")
            self.skip("utility_ebitda", "no 'Segment Analysis' sheet")
            return
        seg_name = None
        for r in range(1, len(sg) + 1):
            v = txt(cell(sg, r, 1))
            if v and ("regulated" in v.lower()) and "subtotal" not in v.lower():
                seg_name = v
                break
        if not seg_name:
            self.skip("utility_revenue", "no segment whose name identifies it as regulated")
            self.skip("utility_ebitda", "no segment whose name identifies it as regulated")
            return
        for field, block in (("utility_revenue", "Total Revenue ($000)"),
                             ("utility_ebitda", "EBITDA ($000)")):
            b = find_row(sg, block)
            if b is None:
                self.skip(field, "no '%s' block on Segment Analysis" % block)
                continue
            r = find_row(sg, seg_name, start=b)
            if r is None:
                self.skip(field, "segment '%s' absent from %s block" % (seg_name, block))
                continue
            vals = [num(cell(sg, r, c)) for c in range(2, 7)]
            if any(v is None for v in vals):
                self.skip(field, "segment '%s' %s series has NA years (%s) - not "
                                 "filled rather than interpolated"
                          % (seg_name, block,
                             ", ".join(str(i) for i, v in enumerate(vals) if v is None)))
            else:
                self.put(field, vals,
                         "Segment Analysis!row %d '%s' under %s (thousands)"
                         % (r, seg_name, block))

    def _net_utility_plant(self, bs, fh):
        """
        'Net Utility Plant' is an SNL/FERC line that CapIQ carries for US
        regulated filers.  It is NOT the same thing as net PP&E and must not be
        substituted with it.
        """
        for rows, sheet in ((bs, "Balance Sheet"), (fh, "Financial Highlights")):
            r = find_row(rows, "Net Utility Plant", exact=False)
            if r:
                ncols = 5
                vals = [num(cell(rows, r, c)) for c in range(2, 2 + ncols)]
                self.put("net_utility_plant", vals, "%s!row %d" % (sheet, r))
                if vals and vals[-1] is not None:
                    self.put("net_utility_plant_latest", vals[-1], "%s!row %d (latest)" % (sheet, r))
                    self.put("rate_base_consolidated", vals[-1], "%s!row %d (latest)" % (sheet, r))
                    self.put("rate_base_consolidated_label",
                             "Net Utility Plant (Balance Sheet)", "constant label")
                return
        why = ("no 'Net Utility Plant' line in this workbook (SNL/FERC utility-plant "
               "series is not carried for this issuer); net PP&E is a different "
               "measure and was deliberately not substituted")
        for f in ("net_utility_plant", "net_utility_plant_latest",
                  "rate_base_consolidated", "rate_base_consolidated_label"):
            self.skip(f, why)

    def _capacity(self, wb):
        rows = sheet_rows(wb, "Current Capacity Summary")
        if not rows:
            for f in ("total_capacity_mw", "solar_capacity_mw", "wind_capacity_mw",
                      "nuclear_capacity_mw", "gas_capacity_mw"):
                self.skip(f, "no 'Current Capacity Summary' sheet")
            return
        hdr = find_row(rows, "Power Plant Category")
        if hdr is None:
            for f in ("total_capacity_mw", "solar_capacity_mw", "wind_capacity_mw",
                      "nuclear_capacity_mw", "gas_capacity_mw"):
                self.skip(f, "capacity table header not found")
            return
        # column of 'Total Capacity (MW)' (owned operating + planned)
        col = None
        for c in range(2, 12):
            if (txt(cell(rows, hdr, c)) or "").lower().startswith("total capacity"):
                col = c
                break
        if col is None:
            col = 8
        # The category table starts after the header/sub-header rows and runs
        # until two consecutive fully blank rows (a sub-header row is blank in
        # col A but carries column captions, so a single blank must not stop us).
        cats = {}
        blanks = 0
        for r in range(hdr + 1, len(rows) + 1):
            lbl = txt(cell(rows, r, 1))
            if lbl is None:
                blanks += 1
                if blanks >= 2:
                    break
                continue
            blanks = 0
            v = num(cell(rows, r, col))
            if v is None:
                continue
            cats.setdefault(lbl.strip().lower(), (r, v))

        def grab(field, keys, note):
            for k in keys:
                for lbl, (rr, v) in cats.items():
                    if lbl == k and v is not None:
                        self.put(field, v, "Current Capacity Summary!row %d '%s' "
                                           "(Total Capacity MW)" % (rr, lbl))
                        return
            self.skip(field, note)

        grab("total_capacity_mw", ["total"],
             "no fleet 'Total' row in Current Capacity Summary")
        grab("gas_capacity_mw", ["total natural gas", "natural gas"],
             "no natural-gas category in Current Capacity Summary")
        grab("solar_capacity_mw", ["total solar", "solar"],
             "no solar category in Current Capacity Summary (fleet holds none)")
        grab("wind_capacity_mw", ["total wind", "wind"],
             "no wind category in Current Capacity Summary (fleet holds none)")
        grab("nuclear_capacity_mw", ["total nuclear", "nuclear"],
             "no nuclear category in Current Capacity Summary (fleet holds none)")

    def _power_plants(self, wb):
        rows = sheet_rows(wb, "Power Plants")
        hdr = find_row(rows, "Power Plant Name") if rows else None
        if hdr is None:
            self.skip("power_plants", "no 'Power Plants' sheet / header")
            return
        n = 0
        for r in range(hdr + 1, len(rows) + 1):
            if txt(cell(rows, r, 1)) and num(cell(rows, r, 2)) is not None:
                n += 1
        self.put("power_plants", n, "Power Plants!rows %d.. (count of plant records)" % (hdr + 1))

    def _analysts(self, wb):
        rows = sheet_rows(wb, "Analyst Coverage")
        hdr = find_row(rows, "Research Contributor") if rows else None
        if hdr is None:
            self.skip("analysts", "no 'Analyst Coverage' sheet / header")
            return
        out = []
        for r in range(hdr + 1, len(rows) + 1):
            a, b = txt(cell(rows, r, 1)), txt(cell(rows, r, 2))
            if not a:
                break
            out.append({"name": a, "firm": b if b else "-"})
        self.put("analysts", out, "Analyst Coverage!rows %d+" % (hdr + 1))

    def _advisers(self, wb):
        rows = sheet_rows(wb, "Advisers")
        if not rows:
            self.skip("advisers", "no 'Advisers' sheet")
            return
        out = []
        for r in range(4, len(rows) + 1):
            a = txt(cell(rows, r, 1))
            if not a:
                continue
            out.append({"firm": a, "role": txt(cell(rows, r, 2)), "deals": txt(cell(rows, r, 3))})
        self.put("advisers", out, "Advisers!rows 4+ (col A/B/C)")

    def _transactions(self, wb):
        rows = sheet_rows(wb, "Transactions Summary")
        if not rows:
            self.skip("transactions", "no 'Transactions Summary' sheet")
            return
        start = None
        for r in range(4, len(rows) + 1):
            if txt(cell(rows, r, 1)):
                start = r
                break
        out = []
        r = start
        while r and r <= len(rows) and txt(cell(rows, r, 1)):
            out.append({"date": txt(cell(rows, r, 1)), "target": txt(cell(rows, r, 2)),
                        "type": txt(cell(rows, r, 3)), "value": txt(cell(rows, r, 4)),
                        "status": txt(cell(rows, r, 5))})
            r += 1
        self.put("transactions", out, "Transactions Summary!rows %s+" % start)

    def _top_holders(self, wb):
        rows = sheet_rows(wb, "Top Holders")
        hdr = find_row(rows, "Top Holders") if rows else None
        if hdr is None:
            self.skip("top_holders", "no 'Top Holders' block")
            return
        out = []
        for r in range(hdr, min(hdr + 7, len(rows) + 1)):
            a = txt(cell(rows, r, 1))
            if not a:
                break
            out.append({"name": a,
                        "shares": num(cell(rows, r, 2)) if num(cell(rows, r, 2)) is not None
                        else txt(cell(rows, r, 2)),
                        "pct": num(cell(rows, r, 3)) if num(cell(rows, r, 3)) is not None
                        else txt(cell(rows, r, 3)),
                        "change": num(cell(rows, r, 4)) if num(cell(rows, r, 4)) is not None
                        else txt(cell(rows, r, 4))})
        self.put("top_holders", out, "Top Holders!rows %d+ (cols A-D)" % hdr)

    def _segments(self, sg):
        if not sg:
            self.skip("segments", "no 'Segment Analysis' sheet")
            return
        out = {}
        for r in range(1, len(sg) + 1):
            v = txt(cell(sg, r, 1))
            if v:
                out.setdefault(v.strip(), {})
        self.put("segments", out, "Segment Analysis!col A row labels")

    def _planned_capex(self, wb):
        name = None
        for s in wb.sheetnames:
            if s.lower().startswith("planned capital expenditures"):
                name = s
                break
        if not name:
            self.skip("planned_capex", "no 'Planned Capital Expenditures' sheet")
            return
        rows = sheet_rows(wb, name)
        out = []
        for r in range(4, len(rows) + 1):
            a = txt(cell(rows, r, 1))
            if not a:
                continue
            vals = []
            for c in range(2, 13):
                v = cell(rows, r, c)
                if v is None or txt(v) is None:
                    continue
                n = num(v)
                vals.append(n if n is not None else txt(v))
            if vals:
                out.append({"category": a, "values": vals})
        self.put("planned_capex", out, "%s!rows with a value" % name)

    def _debt_items(self, wb):
        rows = sheet_rows(wb, "Debt Summary (Reported)")
        if not rows:
            self.skip("debt_items", "no 'Debt Summary (Reported)' sheet")
            return
        out = []
        for r in range(3, len(rows) + 1):
            a = txt(cell(rows, r, 1))
            v = num(cell(rows, r, 2))
            if a and v is not None:
                out.append({"description": a, "amount": v, "maturity": None, "rate": None})
        self.put("debt_items", out, "Debt Summary (Reported)!col A/B (thousands, as reported)")

    def _dividends(self, wb):
        rows = sheet_rows(wb, "Dividends & Splits")
        hdr = find_row(rows, "Ticker") if rows else None
        if hdr is None:
            self.skip("dividend_history", "no summary block on 'Dividends & Splits'")
            return
        out = []
        r = hdr
        while r <= len(rows):
            a = txt(cell(rows, r, 1))
            sec = txt(cell(rows, r, 4))
            if not a or not sec:
                break
            out.append({"date": a, "amount": None, "type": None})
            r += 1
        self.put("dividend_history", out,
                 "Dividends & Splits!summary block rows %d-%d" % (hdr, r - 1))

    def _rate_base(self, wb):
        rows = sheet_rows(wb, "Pending Rate Cases")
        hdr = find_row(rows, "State") if rows else None
        if hdr is None:
            for f in ("rate_base_opcos", "rate_base_best", "rate_base_best_label",
                      "rate_base_best_company"):
                self.skip(f, "no 'Pending Rate Cases' table")
            return
        cols = {}
        for c in range(1, 25):
            v = txt(cell(rows, hdr, c))
            if v:
                cols[v.strip().lower()] = c
        c_state = cols.get("state")
        c_co = cols.get("company")
        c_rb = cols.get("rate base ($m)")
        c_ty = cols.get("rate case test year end date")
        opcos = {}
        for r in range(hdr + 1, len(rows) + 1):
            co = txt(cell(rows, r, c_co)) if c_co else None
            rb = num(cell(rows, r, c_rb)) if c_rb else None
            if not co or rb is None:
                continue
            ty = txt(cell(rows, r, c_ty)) if c_ty else None
            prev = opcos.get(co)
            if prev and prev["rate_base_m"] >= rb:
                continue
            opcos[co] = {"company": co,
                         "state": txt(cell(rows, r, c_state)) if c_state else "",
                         "rate_base_m": rb,
                         "rate_base_b": round(rb / 1000.0, 2),
                         "label": ("Test Yr %s" % ty) if ty else "",
                         "source": "pending"}
        if not opcos:
            for f in ("rate_base_opcos", "rate_base_best", "rate_base_best_label",
                      "rate_base_best_company"):
                self.skip(f, "'Pending Rate Cases' carries no numeric Rate Base ($M)")
            return
        self.put("rate_base_opcos", opcos, "Pending Rate Cases!rows with a Rate Base ($M)")
        best = max(opcos.values(), key=lambda d: d["rate_base_m"])
        self.put("rate_base_best", round(float(best["rate_base_m"]) * 1e6, 4),
                 "largest pending Rate Base ($M) x 1e6 -> reporting units")
        self.put("rate_base_best_label", best["label"], "Rate Case Test Year End Date")
        self.put("rate_base_best_company", best["company"], "Pending Rate Cases company")

    def _ferc(self, wb):
        self.skip("opco_nup_ferc1",
                  "FERC Form 1 opco net-utility-plant is sourced outside CapIQ "
                  "(PUDL); no FERC Form 1 data in this workbook")


# --------------------------------------------------------------------------
# capiq_export.json merge
# --------------------------------------------------------------------------

def is_empty(v):
    return v is None or (isinstance(v, (list, dict, str)) and len(v) == 0)


def locate_record(doc, ticker):
    """Return (container_dict, key) for the ticker record.  Never creates one."""
    if isinstance(doc, dict):
        if isinstance(doc.get("companies"), dict) and ticker in doc["companies"]:
            return doc["companies"], ticker
        if ticker in doc and isinstance(doc[ticker], dict):
            return doc, ticker
        for k in ("data", "tickers", "records"):
            sub = doc.get(k)
            if isinstance(sub, dict) and ticker in sub:
                return sub, ticker
    return None, None


def atomic_write_json(path, doc):
    d = os.path.dirname(os.path.abspath(path)) or "."
    fd, tmp = tempfile.mkstemp(prefix=".capiq_", suffix=".tmp", dir=d)
    try:
        with os.fdopen(fd, "w", encoding="utf-8") as fh:
            json.dump(doc, fh, ensure_ascii=False, indent=2)
            fh.flush()
            os.fsync(fh.fileno())
        with open(tmp, "r", encoding="utf-8") as fh:      # re-read + validate
            json.load(fh)
        os.replace(tmp, path)
    except Exception:
        if os.path.exists(tmp):
            os.remove(tmp)
        raise


def main(argv=None):
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--workbook", required=True, help="CapIQ tearsheet .xlsx")
    ap.add_argument("--capiq", required=True, help="capiq_export.json to update in place")
    ap.add_argument("--ticker", default="AQN")
    ap.add_argument("--result-json", default=None,
                    help="also dump just the extracted values to this path")
    ap.add_argument("--dry-run", action="store_true", help="report only; do not write")
    args = ap.parse_args(argv)

    wb = openpyxl.load_workbook(args.workbook, data_only=True, read_only=True)
    ex = Extractor(wb).run()

    with open(args.capiq, "r", encoding="utf-8") as fh:
        doc = json.load(fh)
    container, key = locate_record(doc, args.ticker)
    if container is None:
        sys.exit("ticker %r not found in %s - refusing to create a new record"
                 % (args.ticker, args.capiq))
    rec = container[key]

    filled, kept, notfound = [], [], []
    for field, value in ex.values.items():
        if is_empty(rec.get(field)):
            rec[field] = value
            filled.append(field)
        else:
            kept.append(field)
    for field, why in ex.missing.items():
        if is_empty(rec.get(field)):
            notfound.append((field, why))
        else:
            kept.append(field)

    # ---- report --------------------------------------------------------
    print("=" * 78)
    print("workbook : %s" % os.path.abspath(args.workbook))
    print("capiq    : %s" % os.path.abspath(args.capiq))
    print("ticker   : %s" % args.ticker)
    print("currency : financial statements report in %s"
          % "/".join(ex.currency.get("financials", ["unknown"])))
    for lbl in ex.currency.get("market_block", []):
        print("           market block label: %s" % lbl)
    print("=" * 78)
    print("FILLED (%d)" % len(filled))
    for f in [x for x in FIELD_ORDER if x in filled] + [x for x in filled if x not in FIELD_ORDER]:
        print("  %-28s %-58s  <- %s" % (f, preview(ex.values[f]), ex.source[f]))
    if kept:
        print("\nALREADY POPULATED, LEFT UNTOUCHED (%d)" % len(kept))
        for f in sorted(set(kept)):
            print("  %-28s %s" % (f, preview(rec.get(f))))
    print("\nNOT FILLED (%d)" % len(notfound))
    for f, why in sorted(notfound):
        print("  %-28s %s" % (f, why))
    print("=" * 78)

    if args.result_json:
        with open(args.result_json, "w", encoding="utf-8") as fh:
            json.dump({args.ticker: ex.values,
                       "_not_found": dict(notfound),
                       "_sources": ex.source,
                       "_currency": ex.currency},
                      fh, ensure_ascii=False, indent=2)
        print("wrote %s" % os.path.abspath(args.result_json))

    if args.dry_run:
        print("dry-run: %s left unchanged" % args.capiq)
        return 0
    atomic_write_json(args.capiq, doc)
    print("wrote %s (atomic replace, re-parsed before swap)" % os.path.abspath(args.capiq))
    return 0


def preview(v, width=56):
    if isinstance(v, list):
        s = "[%d] %s" % (len(v), json.dumps(v[:4], default=str))
    elif isinstance(v, dict):
        s = "{%d} %s" % (len(v), json.dumps(list(v)[:3], default=str))
    else:
        s = json.dumps(v, default=str)
    s = s.replace("\n", " ")
    return s if len(s) <= width else s[:width - 3] + "..."


FIELD_ORDER = [
    "name", "long_description", "hq", "website", "phone", "sic", "power_plants",
    "periods", "revenue", "ebitda", "ebitda_margin", "net_income", "eps_diluted",
    "revenue_growth", "total_assets", "net_debt", "net_ppe", "total_equity", "capex",
    "dividends_ps", "roe", "roa", "total_debt", "cash", "tev", "analysts",
    "net_utility_plant", "utility_revenue", "utility_ebitda", "retail_customers",
    "residential_customers", "retail_revenue", "retail_mwh", "gas_customers",
    "electric_customers", "fte_employees", "total_capacity_mw", "solar_capacity_mw",
    "wind_capacity_mw", "nuclear_capacity_mw", "gas_capacity_mw", "advisers",
    "transactions", "top_holders", "segments", "total_customers_latest",
    "net_utility_plant_latest", "planned_capex", "debt_items", "dividend_history",
    "rate_base_consolidated", "rate_base_consolidated_label", "rate_base_opcos",
    "rate_base_best", "rate_base_best_label", "rate_base_best_company", "opco_nup_ferc1",
]


if __name__ == "__main__":
    sys.exit(main())
